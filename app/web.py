"""FastAPI app: static web UI + JSON API. Served by gunicorn/uvicorn (see Dockerfile)."""
import hashlib
import io
import json
import os
import secrets
import urllib.request
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from . import config, db, kis, ledger, movements, trading, valuation
from . import market as market_mod
from .realestate.seoul import SEOUL_GU

app = FastAPI(title="자산현황")

STATIC = Path(__file__).resolve().parent / "static"
app.mount("/static", StaticFiles(directory=STATIC), name="static")

AUTH_URL = os.environ.get("AUTH_URL", "http://auth-app:8000")


def _conn():
    return db.connect()


@app.get("/health")
def health():
    with _conn() as conn:
        conn.execute("SELECT 1")
    return {"status": "ok"}


@app.get("/")
def index():
    return FileResponse(STATIC / "index.html")


def _current_user(request: Request):
    """로그인 유저(auth-server /api/me 프록시). 없으면 None."""
    sid = request.cookies.get("sid")
    if not sid:
        return None
    try:
        req = urllib.request.Request(f"{AUTH_URL}/api/me", headers={"Cookie": f"sid={sid}"})
        with urllib.request.urlopen(req, timeout=5) as r:
            d = json.loads(r.read())
        return d.get("user") if d.get("authenticated") else None
    except Exception:
        return None


def _require_admin(request: Request):
    u = _current_user(request)
    return u if (u and u.get("role") == "admin") else None


def _auth_get(path: str, request: Request):
    sid = request.cookies.get("sid") or ""
    req = urllib.request.Request(f"{AUTH_URL}{path}", headers={"Cookie": f"sid={sid}"})
    with urllib.request.urlopen(req, timeout=6) as r:
        return json.loads(r.read())


def _auth_post(path: str, payload: dict, request: Request):
    sid = request.cookies.get("sid") or ""
    data = json.dumps(payload).encode()
    req = urllib.request.Request(f"{AUTH_URL}{path}", data=data, method="POST",
                                 headers={"Cookie": f"sid={sid}", "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=6) as r:
        return json.loads(r.read())


@app.get("/api/whoami")
def api_whoami(request: Request):
    """우상단 프로필 표시용."""
    u = _current_user(request)
    return {"authenticated": True, "user": u} if u else {"authenticated": False}


@app.get("/api/kis/status")
def api_kis_status():
    """자동매매(KIS) 설정 상태(현재 env 기준). 자격증명은 마스킹."""
    env = kis._env()
    acc = kis._creds(env)[2] or ""
    masked = (acc[:4] + "***" + acc[-2:]) if len(acc) >= 6 else ("설정됨" if acc else "")
    return {"configured": kis.configured(env), "env": env, "account": masked,
            "base": kis._base(env), "live_allowed": config.KIS_ALLOW_LIVE,
            "vts_configured": kis.configured("vts"), "prod_configured": kis.configured("prod")}


class KisEnvIn(BaseModel):
    env: str   # vts | prod


@app.post("/api/kis/env")
def api_kis_env(body: KisEnvIn, request: Request):
    """모의(vts)/실전(prod) 전환. 관리자 전용. 실전은 별도 자격증명·KIS_ALLOW_LIVE 필요."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자만 전환할 수 있습니다."}, status_code=403)
    try:
        return {"ok": True, "env": kis.set_env(body.env)}
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/tax")
def api_tax():
    """세금 추정(참고): 연도별 국내/해외 실현손익 + 배당·이자, 해외주식 양도세·금융소득 종합과세 여부.
    ⚠️ 추정치 — 매도 시점 FX 대신 현재 FX 사용, 국내주식 양도세(대주주)·건보료 미반영."""
    with _conn() as conn:
        t = movements.tax_by_year(conn)
        fxrow = conn.execute("SELECT price FROM prices WHERE price_key = %s", ("FX:USDKRW",)).fetchone()
    fx = float((fxrow and fxrow["price"]) or 1300.0)
    years = sorted(set(list(t["realized"].keys()) + list(t["divint"].keys())), reverse=True)
    rows = []
    for y in years:
        r, di = t["realized"].get(y, {}), t["divint"].get(y, {})
        dom = r.get("KRW", 0.0)                       # 국내주식 실현손익(원)
        foreign = r.get("USD", 0.0) * fx              # 해외주식 실현손익(원화환산)
        foreign_cgt = max(0.0, foreign - 2_500_000) * 0.22   # 해외 양도세: (이익-250만)×22%
        fin_income = di.get("KRW", 0.0) + di.get("USD", 0.0) * fx   # 금융소득(배당+이자)
        rows.append({"year": y, "domestic_realized": round(dom), "foreign_realized": round(foreign),
                     "foreign_cgt": round(foreign_cgt), "fin_income": round(fin_income),
                     "comprehensive": fin_income > 20_000_000})
    return {"fx": round(fx, 2), "rows": rows}


@app.get("/api/kis/balance")
def api_kis_balance():
    """자동매매(KIS) 국내주식 잔고(보유종목·예수금)."""
    try:
        return kis.balance()
    except Exception as e:
        return {"error": str(e)}


class KisOrderIn(BaseModel):
    symbol: str
    qty: int
    side: str = "buy"       # buy | sell
    market: bool = True     # True=시장가, False=지정가
    price: int = 0          # 지정가일 때 단가


@app.post("/api/kis/order")
def api_kis_order(o: KisOrderIn, request: Request):
    """자동매매(KIS) 현금 주문(관리자 전용). 실전 실주문은 KIS_ALLOW_LIVE로 별도 보호."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자만 주문할 수 있습니다."}, status_code=403)
    try:
        return kis.order(o.symbol, o.qty, o.side, o.price, o.market)
    except Exception as e:
        return {"error": str(e)}


# ── 단타 규칙 엔진(모의투자 vts 기본): 이동평균±k×ATR 밴드 평균회귀 ──────
class TradeRuleIn(BaseModel):
    id: Optional[int] = None
    symbol: str
    name: Optional[str] = None
    strategy: str = "band"          # band(밴드) | grid(사다리)
    timeframe: str = "intraday"     # band: intraday(장중분) | daily(일봉스윙)
    ma_window: int = 20             # band: 관찰 분(장중) 또는 일수(일봉)
    vol_mult: float = 1.5           # band: k
    grid_step: int = 100            # grid: 간격(원)
    grid_levels: int = 5            # grid: 단계
    center: Optional[int] = None    # grid: 기준가(없으면 최초 현재가)
    max_position: int = 0           # grid/bandgrid: 보유 상한(주, 0=무제한)
    qty: int = 1
    env: str = "vts"
    order_type: str = "market"      # market(시장가) | ioc(IOC지정가·현재가 지정, 미체결 자동취소)
    gap_ticks: int = 2              # custom: 층 간격(호가단위 배수)
    cash_share: float = 0.10        # custom: 층당 예수금 비중
    eod_ratio: float = 0.0          # custom: 종가 정리 비율(0=안 함)


@app.get("/api/trade/rules")
def api_trade_rules():
    with _conn() as conn:
        db.init_schema(conn)
        rules = [dict(r) for r in conn.execute("SELECT * FROM trade_rules ORDER BY id")]
    for r in rules:
        # bandgrid는 매도를 밴드가 아니라 '산 값 + step'으로 한다 → 실제 목표가를 따로 준다.
        r["sell_targets"] = []
        if r.get("strategy") == "bandgrid":
            try:
                lots = json.loads(r.get("state") or "{}").get("lots", [])
            except Exception:
                lots = []
            step = int(r.get("grid_step") or 0)
            r["sell_targets"] = sorted(int(l["buy"]) + step for l in lots if l.get("buy"))
    return {"rules": rules, "market_open": trading.market_open()}


@app.post("/api/trade/rules")
def api_trade_rule_save(body: TradeRuleIn, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    env = "vts" if body.env != "prod" else "prod"
    strat = body.strategy if body.strategy in ("grid", "bandgrid", "custom") else "band"
    tf = "daily" if body.timeframe == "daily" else "intraday"
    otype = body.order_type if body.order_type in ("market", "ioc") else "market"
    cols = (body.symbol, body.name, strat, tf, body.ma_window, body.vol_mult,
            body.grid_step, body.grid_levels, body.center, body.max_position, body.qty, env, otype,
            max(1, body.gap_ticks), min(1.0, max(0.01, body.cash_share)), min(1.0, max(0.0, body.eod_ratio)))
    with _conn() as conn:
        db.init_schema(conn)
        if body.id:
            conn.execute("""UPDATE trade_rules SET symbol=%s,name=%s,strategy=%s,timeframe=%s,ma_window=%s,vol_mult=%s,
                            grid_step=%s,grid_levels=%s,center=%s,max_position=%s,qty=%s,env=%s,order_type=%s,
                            gap_ticks=%s,cash_share=%s,eod_ratio=%s,
                            state=NULL,ticks=NULL,ma=NULL,atr=NULL,band_buy=NULL,band_sell=NULL,
                            last_price=NULL,last_eval=NULL,base_cash=0 WHERE id=%s""",
                         (*cols, body.id))
            rid = body.id
        else:
            rid = conn.execute("""INSERT INTO trade_rules(symbol,name,strategy,timeframe,ma_window,vol_mult,grid_step,grid_levels,center,max_position,qty,env,order_type,gap_ticks,cash_share,eod_ratio)
                                  VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""", cols).fetchone()["id"]
        conn.commit()
    return {"ok": True, "id": rid}


@app.post("/api/trade/rules/{rid}/toggle")
def api_trade_rule_toggle(rid: int, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("UPDATE trade_rules SET active = NOT active WHERE id=%s", (rid,))
        conn.commit()
        r = conn.execute("SELECT active FROM trade_rules WHERE id=%s", (rid,)).fetchone()
    return {"ok": True, "active": r["active"] if r else None}


@app.delete("/api/trade/rules/{rid}")
def api_trade_rule_del(rid: int, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("DELETE FROM trade_rules WHERE id=%s", (rid,))
        conn.commit()
    return {"ok": True}


@app.post("/api/trade/tick")
def api_trade_tick(request: Request):
    """지금 즉시 활성 규칙 평가(관리자, 장 마감이어도 강제). 조건 충족 시 vts 주문."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        db.init_schema(conn)
        return trading.tick(conn, force=True)


@app.get("/api/trade/log")
def api_trade_log(date_from: Optional[str] = None, date_to: Optional[str] = None, limit: int = 200):
    """체결 로그(기간 필터) + 일자별·기간 손익.

    실현손익은 전체 이력을 FIFO로 훑으며 '그 기간에 팔린 것'의 손익만 집계한다
    (기간 시작 전에 산 물량을 기간 중에 팔았을 때 취득원가를 제대로 쓰기 위해).
    비용(수수료·거래세)은 체결 시점에 기록된 값을 그대로 더한다."""
    from collections import deque
    with _conn() as conn:
        db.init_schema(conn)
        allrows = conn.execute(
            "SELECT ts, side, qty, price, fee, tax FROM trade_log ORDER BY ts, id").fetchall()
        where, params = [], []
        if date_from:
            where.append("ts >= %s"); params.append(date_from)
        if date_to:
            where.append("ts <= %s"); params.append(date_to + " 23:59:59")
        w = ("WHERE " + " AND ".join(where)) if where else ""
        rows = [dict(r) for r in conn.execute(
            f"SELECT * FROM trade_log {w} ORDER BY id DESC LIMIT %s", (*params, limit))]
        total = conn.execute(f"SELECT count(*) n FROM trade_log {w}", tuple(params)).fetchone()["n"]
        dates = [r["ts"] for r in conn.execute("SELECT DISTINCT left(ts,10) ts FROM trade_log ORDER BY 1 DESC")]
        px = {r["symbol"]: r["last_price"] for r in
              conn.execute("SELECT symbol, last_price FROM trade_rules").fetchall()}
        sym = conn.execute("SELECT symbol FROM trade_log ORDER BY id DESC LIMIT 1").fetchone()

    def inrange(d):
        return (not date_from or d >= date_from) and (not date_to or d <= date_to)

    lots = deque()
    days, tot = {}, {"buy_amt": 0.0, "sell_amt": 0.0, "buy_qty": 0.0, "sell_qty": 0.0,
                     "fee": 0.0, "tax": 0.0, "realized": 0.0}
    all_realized = all_cost = 0.0
    for r in allrows:
        d = str(r["ts"])[:10]
        q, p = r["qty"] or 0, r["price"] or 0
        fee, tax = r["fee"] or 0, r["tax"] or 0
        all_cost += fee + tax
        acc = None
        if inrange(d):
            acc = days.setdefault(d, {"date": d, "buy_amt": 0.0, "sell_amt": 0.0, "buy_qty": 0.0,
                                      "sell_qty": 0.0, "fee": 0.0, "tax": 0.0, "realized": 0.0,
                                      "open_qty": 0.0})
            acc["fee"] += fee; acc["tax"] += tax
            tot["fee"] += fee; tot["tax"] += tax
            acc["_open"] = True
        if r["side"] == "buy":
            lots.append([q, p])
            if acc is not None:
                acc["buy_amt"] += q * p; acc["buy_qty"] += q
                tot["buy_amt"] += q * p; tot["buy_qty"] += q
                acc["open_qty"] = sum(l[0] for l in lots)   # 그 시점 잔여 → 마지막 체결 값이 그날 마감 잔여
            continue
        rem, pnl = q, 0.0
        while rem > 0 and lots:
            lot = lots[0]
            take = min(rem, lot[0])
            pnl += (p - lot[1]) * take
            lot[0] -= take; rem -= take
            if lot[0] == 0:
                lots.popleft()
        all_realized += pnl
        if acc is not None:
            acc["sell_amt"] += q * p; acc["sell_qty"] += q; acc["realized"] += pnl
            acc["open_qty"] = sum(l[0] for l in lots)
            tot["sell_amt"] += q * p; tot["sell_qty"] += q; tot["realized"] += pnl

    openq = sum(l[0] for l in lots)
    opencost = sum(l[0] * l[1] for l in lots)
    cur = px.get(sym["symbol"]) if sym else None
    unreal = (openq * cur - opencost) if (cur and openq) else 0.0
    rnd = lambda d: {k: (round(v) if isinstance(v, float) else v) for k, v in d.items()}
    for v in days.values():
        v["net"] = v["realized"] - v["fee"] - v["tax"]
        v.pop("_open", None)
    tot["net"] = tot["realized"] - tot["fee"] - tot["tax"]
    return {
        "log": rows, "total": total, "dates": dates,
        "period": {**rnd(tot), "fills": len(rows), "open_qty": openq},
        "days": [rnd(v) for v in sorted(days.values(), key=lambda x: x["date"], reverse=True)],
        "summary": {"fills": len(allrows), "realized": round(all_realized), "unrealized": round(unreal),
                    "cost": round(all_cost), "net": round(all_realized + unreal - all_cost),
                    "open_qty": openq, "open_avg": round(opencost / openq) if openq else 0,
                    "cur_price": cur}}


@app.get("/api/trade/chart")
def api_trade_chart(symbol: str, ma: int = 20, k: float = 1.5, days: int = 120):
    """종목 최근 일봉 + MA + 밴드(MA±k×ATR) 시계열 — 매매 페이지 차트용."""
    try:
        from datetime import date as _date
        import FinanceDataReader as fdr
        import pandas as pd
        df = fdr.DataReader(symbol, str(_date.today().year - 2) + "-01-01")
        if df is None or len(df) < ma + 15:
            return {"error": "시세 데이터 부족"}
        c, h, l = df["Close"], df["High"], df["Low"]
        mav = c.rolling(ma).mean()
        tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
        atr = tr.rolling(14).mean()
        buy, sell = mav - k * atr, mav + k * atr
        nn = lambda x: None if pd.isna(x) else float(x)
        bars = [{"date": str(d.date()), "close": float(c[d]), "ma": nn(mav[d]),
                 "buy": nn(buy[d]), "sell": nn(sell[d])} for d in df.tail(days).index]
        return {"symbol": symbol, "bars": bars, "last": float(c.iloc[-1]),
                "ma": float(mav.iloc[-1]), "atr": float(atr.iloc[-1]),
                "buy": float(buy.iloc[-1]), "sell": float(sell.iloc[-1])}
    except Exception as e:
        return {"error": str(e)}


# ── 문서정리(보험 등): 업로드 → OCR/텍스트추출 → 키워드 분류 → 보관 ──────
_DOC_COLS = ("id, orig_name, mime, status, category, person, hospital, doc_type, diagnosis, "
             "amount, doc_date, date_end, claim_group, claimed, claim_ins, sort_order, "
             "insurer, policy_no, expiry_date, uploaded_at")


@app.get("/api/family")
def api_family(request: Request):
    """가족 명단 + 로그인 사용자(auth). 관리자 전용. 최초엔 owners로 가족 시드."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        if not conn.execute("SELECT 1 FROM family LIMIT 1").fetchone():
            conn.execute("INSERT INTO family(name) SELECT DISTINCT name FROM owners ON CONFLICT DO NOTHING")
            conn.commit()
        fam = conn.execute("SELECT id, name, relation, note FROM family ORDER BY id").fetchall()
        owners = conn.execute("SELECT name, include_totals FROM owners ORDER BY name").fetchall()
    try:
        users = _auth_get("/api/users", request).get("users", [])
    except Exception:
        users = []
    return {"family": fam, "users": users, "owners": owners}


class OwnerInclude(BaseModel):
    owner: str
    include: bool = True


@app.post("/api/owners/include")
def api_owner_include(body: OwnerInclude, request: Request):
    """소유자 자산 집계 포함/제외 토글(설정>가족). 대시보드/자산/투자 집계에 반영."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("UPDATE owners SET include_totals=%s WHERE name=%s", (bool(body.include), body.owner))
        conn.commit()
    return {"ok": True}


class FamilyIn(BaseModel):
    name: str
    relation: Optional[str] = None
    note: Optional[str] = None


@app.post("/api/family")
def api_family_add(body: FamilyIn, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    name = (body.name or "").strip()
    if not name:
        return JSONResponse({"error": "이름을 입력하세요"}, status_code=400)
    with _conn() as conn:
        conn.execute("""INSERT INTO family(name, relation, note) VALUES(%s,%s,%s)
                        ON CONFLICT (name) DO UPDATE SET relation=EXCLUDED.relation, note=EXCLUDED.note""",
                     (name, body.relation, body.note))
        conn.commit()
    return {"ok": True}


@app.patch("/api/family/{fid}")
def api_family_edit(fid: int, body: FamilyIn, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("UPDATE family SET name=%s, relation=%s, note=%s WHERE id=%s",
                     ((body.name or "").strip(), body.relation, body.note, fid))
        conn.commit()
    return {"ok": True}


@app.delete("/api/family/{fid}")
def api_family_del(fid: int, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("DELETE FROM family WHERE id=%s", (fid,))
        conn.commit()
    return {"ok": True}


class FamilyUserAction(BaseModel):
    id: int
    op: str                       # approve | revoke | owner | delete
    owner: Optional[str] = None


@app.post("/api/family/user-action")
def api_family_user_action(body: FamilyUserAction, request: Request):
    """로그인 사용자 승인/해제/소유자지정/삭제 → auth-server 프록시."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    try:
        return _auth_post("/api/users/action", body.dict(), request)
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/portfolio")
def api_portfolio():
    with _conn() as conn:
        return valuation.portfolio(conn)


@app.get("/api/accounts")
def api_accounts():
    with _conn() as conn:
        return [dict(a) for a in ledger.all_accounts(conn)]


@app.get("/api/cashbook-summary")
def api_cashbook_summary(owner: Optional[str] = None, account_id: Optional[str] = None,
                        date_from: Optional[str] = None, date_to: Optional[str] = None):
    filt, fp = [], []
    if _multi(owner):
        filt.append("o.name = ANY(%s)"); fp.append(_multi(owner))
    if _multi_int(account_id):
        filt.append("a.id = ANY(%s)"); fp.append(_multi_int(account_id))
    if date_from:
        filt.append("t.trade_date >= %s"); fp.append(date_from)
    if date_to:
        filt.append("t.trade_date <= %s"); fp.append(date_to)
    fclause = ("" if not filt else " AND " + " AND ".join(filt))
    joins = ("""FROM transactions t
                JOIN accounts a ON a.id = t.account_id
                JOIN owners o ON o.id = a.owner_id
                LEFT JOIN prices p ON t.currency <> 'KRW' AND p.price_key = 'FX:' || t.currency || 'KRW'""")
    with _conn() as conn:
        # 현금거래 + 배당(수입) 유형별 합계
        rows = conn.execute(
            f"""SELECT t.type, sum(t.amount * COALESCE(p.price, 1)) AS krw {joins}
                WHERE t.type = ANY(%s){fclause} GROUP BY t.type""",
            [list(_CASHFLOW) + ["DIVIDEND"]] + fp,
        ).fetchall()
        # 매매 정산액(중립): 매수 유출 / 매도 유입 — 순증감엔 미포함, 참고용
        tr = conn.execute(
            f"""SELECT
                  COALESCE(sum((t.quantity*t.price + t.fee + t.tax) * COALESCE(p.price,1))
                           FILTER (WHERE t.type='BUY'), 0) AS buy,
                  COALESCE(sum((t.quantity*t.price - t.fee - t.tax) * COALESCE(p.price,1))
                           FILTER (WHERE t.type='SELL'), 0) AS sell {joins}
                WHERE t.type = ANY(ARRAY['BUY','SELL']){fclause}""",
            fp,
        ).fetchone()
    by = {r["type"]: float(r["krw"] or 0) for r in rows}
    income = by.get("DEPOSIT", 0) + by.get("INTEREST", 0) + by.get("DIVIDEND", 0)
    expense = by.get("WITHDRAWAL", 0) + by.get("FEE", 0) + by.get("TAX", 0)
    buy, sell = float(tr["buy"] or 0), float(tr["sell"] or 0)
    return {"income": income, "expense": expense, "net": income - expense,
            "transfer": by.get("TRANSFER", 0), "exchange": by.get("EXCHANGE", 0),
            "dividend": by.get("DIVIDEND", 0), "buy": buy, "sell": sell, "trade_net": sell - buy,
            "by": by}


@app.get("/api/meta")
def api_meta():
    """Filter options for the transaction ledger."""
    with _conn() as conn:
        owners = [r["name"] for r in conn.execute("SELECT name FROM owners ORDER BY name")]
        accounts = [dict(a) for a in ledger.all_accounts(conn)]
        types = [r["type"] for r in
                 conn.execute("SELECT DISTINCT type FROM transactions ORDER BY type")]
    return {"owners": owners, "accounts": accounts, "types": types}


_TX_SORT = {
    "trade_date": "t.trade_date", "type": "t.type", "quantity": "t.quantity",
    "price": "t.price", "owner": "o.name", "name": "COALESCE(t.name, t.symbol)",
    "amount": "(CASE WHEN t.type IN ('BUY','SELL') THEN t.price*t.quantity ELSE t.amount END)",
}


def _multi(v):
    """콤마 구분 문자열 → 리스트(빈 값이면 None)."""
    if not v:
        return None
    out = [s.strip() for s in str(v).split(",") if s.strip()]
    return out or None


def _multi_int(v):
    if not v:
        return None
    out = [int(s) for s in str(v).split(",") if s.strip().lstrip("-").isdigit()]
    return out or None


_INVEST = ("BUY", "SELL", "DIVIDEND", "TRANSFER_IN", "TRANSFER_OUT")
_CASHFLOW = ("DEPOSIT", "WITHDRAWAL", "INTEREST", "FEE", "TAX",
             "XFER_IN", "XFER_OUT", "FX_IN", "FX_OUT", "IPO_IN", "IPO_OUT",
             "TRANSFER", "EXCHANGE", "SUBSCRIPTION")  # 뒤 3개는 구 데이터 하위호환
_QTY_TYPES = ("BUY", "SELL", "TRANSFER_IN", "TRANSFER_OUT")   # 수량×단가 있는 유형


@app.get("/api/transactions")
def api_transactions(
    owner: Optional[str] = None,
    account_id: Optional[str] = None,
    type: Optional[str] = None,
    category: Optional[str] = None,   # invest | cashbook
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: str = "trade_date",
    dir: str = "desc",
    limit: int = 100,
    offset: int = 0,
):
    clauses, params = [], []
    if category == "invest":
        clauses.append("t.type = ANY(%s)"); params.append(list(_INVEST))
    elif category == "cashbook":
        clauses.append("t.type = ANY(%s)"); params.append(list(_CASHFLOW))
    if _multi(owner):
        clauses.append("o.name = ANY(%s)"); params.append(_multi(owner))
    if _multi_int(account_id):
        clauses.append("a.id = ANY(%s)"); params.append(_multi_int(account_id))
    if type:
        clauses.append("t.type = %s"); params.append(type)
    if q:
        clauses.append("(t.symbol ILIKE %s OR t.name ILIKE %s)")
        params += [f"%{q}%", f"%{q}%"]
    if date_from:
        clauses.append("t.trade_date >= %s"); params.append(date_from)
    if date_to:
        clauses.append("t.trade_date <= %s"); params.append(date_to)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    joins = ("FROM transactions t JOIN accounts a ON a.id = t.account_id "
             "JOIN owners o ON o.id = a.owner_id")
    order_col = _TX_SORT.get(sort, "t.trade_date")
    order_dir = "ASC" if dir == "asc" else "DESC"
    with _conn() as conn:
        total = conn.execute(f"SELECT count(*) AS c {joins} {where}", params).fetchone()["c"]
        rows = conn.execute(
            f"""SELECT t.id, t.trade_date, t.type, t.symbol, t.name, t.currency,
                       t.quantity, t.price, t.amount, t.fee, t.tax,
                       a.brokerage, a.account_no, a.alias, o.name AS owner_name
                {joins} {where}
                ORDER BY {order_col} {order_dir}, t.id DESC LIMIT %s OFFSET %s""",
            params + [limit, offset],
        ).fetchall()
    return {"total": total, "rows": rows, "limit": limit, "offset": offset}


@app.get("/api/invest")
def api_invest(
    owner: Optional[str] = None,
    account_id: Optional[str] = None,
    type: Optional[str] = None,     # BUY|SELL|DIVIDEND
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: str = "trade_date",
    dir: str = "desc",
    limit: int = 100,
    offset: int = 0,
):
    """투자내역: (날짜·계좌·종목·유형) 요약 그룹 + 각 그룹의 체결(fills)."""
    from collections import defaultdict
    clauses, params = ["t.type = ANY(%s)"], [list(_INVEST)]
    if type in _INVEST:
        clauses, params = ["t.type = %s"], [type]
    if _multi(owner):
        clauses.append("o.name = ANY(%s)"); params.append(_multi(owner))
    if _multi_int(account_id):
        clauses.append("a.id = ANY(%s)"); params.append(_multi_int(account_id))
    if q:
        clauses.append("(t.symbol ILIKE %s OR t.name ILIKE %s)"); params += [f"%{q}%", f"%{q}%"]
    if date_from:
        clauses.append("t.trade_date >= %s"); params.append(date_from)
    if date_to:
        clauses.append("t.trade_date <= %s"); params.append(date_to)
    where = "WHERE " + " AND ".join(clauses)
    with _conn() as conn:
        fills = conn.execute(
            f"""SELECT t.id, t.trade_date, t.type, t.symbol, t.name, t.market, t.currency,
                       t.quantity, t.price, t.amount, t.fee, t.tax, t.source, t.src_row,
                       a.id AS account_id, a.brokerage, a.account_no, a.alias, o.name AS owner_name
                FROM transactions t JOIN accounts a ON a.id = t.account_id
                JOIN owners o ON o.id = a.owner_id
                {where} ORDER BY t.trade_date DESC, t.id""",
            params).fetchall()
    # 중복 의심: 동일 (계좌·날짜·유형·종목·수량·단가)가 서로 다른 source에서 오면 표시
    srcmap = defaultdict(set)
    for f in fills:
        srcmap[(f["account_id"], f["trade_date"], f["type"], f["symbol"], f["quantity"], f["price"])].add(f["source"] or "수동")
    groups, order = {}, []
    for f in fills:
        key = (f["trade_date"], f["account_id"], f["symbol"], f["type"])
        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                "trade_date": f["trade_date"], "type": f["type"], "symbol": f["symbol"],
                "name": f["name"], "market": f["market"], "currency": f["currency"],
                "owner_name": f["owner_name"], "brokerage": f["brokerage"],
                "alias": f["alias"], "account_no": f["account_no"], "account_id": f["account_id"],
                "quantity": 0.0, "_pv": 0.0, "amount": 0.0, "fee": 0.0, "tax": 0.0,
                "fills": [], "dup_suspect": False,
            }
            order.append(key)
        g["quantity"] += f["quantity"] or 0
        g["_pv"] += (f["quantity"] or 0) * (f["price"] or 0)
        g["amount"] += f["amount"] or 0
        g["fee"] += f["fee"] or 0
        g["tax"] += f["tax"] or 0
        dup = len(srcmap[(f["account_id"], f["trade_date"], f["type"], f["symbol"], f["quantity"], f["price"])]) > 1
        g["dup_suspect"] = g["dup_suspect"] or dup
        g["fills"].append({"id": f["id"], "quantity": f["quantity"], "price": f["price"],
                           "amount": f["amount"], "fee": f["fee"], "tax": f["tax"],
                           "source": f["source"], "src_row": f["src_row"], "dup_suspect": dup})
    out = []
    for key in order:
        g = groups[key]
        g["avg_price"] = (g["_pv"] / g["quantity"]) if g["quantity"] else None
        g["value"] = (g["quantity"] * g["avg_price"]) if (g["type"] in _QTY_TYPES and g["avg_price"]) else g["amount"]
        del g["_pv"]
        out.append(g)
    # 동일 날짜 안에서도 체결 id(입력순)로 정렬되도록 보조키 사용
    _mid = lambda g: max((f["id"] for f in g["fills"]), default=0)
    keyfn = {"trade_date": lambda g: (g["trade_date"], _mid(g)),
             "owner": lambda g: g["owner_name"] or "",
             "symbol": lambda g: g["name"] or g["symbol"] or "",
             "value": lambda g: g["value"] or 0,
             "type": lambda g: g["type"]}.get(sort, lambda g: (g["trade_date"], _mid(g)))
    out.sort(key=keyfn, reverse=(dir != "asc"))
    return {"total": len(out), "groups": out[offset:offset + limit], "limit": limit, "offset": offset}


_CASH_INCOME = ("DEPOSIT", "INTEREST")
_CASH_EXPENSE = ("WITHDRAWAL", "FEE", "TAX")


@app.get("/api/cashbook")
def api_cashbook(
    owner: Optional[str] = None,
    account_id: Optional[str] = None,
    type: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: str = "trade_date",
    dir: str = "desc",
    limit: int = 100,
    offset: int = 0,
):
    """통합 현금원장: 순수 현금거래(입출금·이체·환전·이자·수수료·세금) + 투자 정산 레그
    (매수 −정산액, 매도 +정산액, 배당 +). 수입/지출/중립 부호. 매매는 중립(자산 재배치)."""
    def _filt(clauses, params):
        if _multi(owner):
            clauses.append("o.name = ANY(%s)"); params.append(_multi(owner))
        if _multi_int(account_id):
            clauses.append("a.id = ANY(%s)"); params.append(_multi_int(account_id))
        if q:
            clauses.append("(t.symbol ILIKE %s OR t.name ILIKE %s)"); params += [f"%{q}%", f"%{q}%"]
        if date_from:
            clauses.append("t.trade_date >= %s"); params.append(date_from)
        if date_to:
            clauses.append("t.trade_date <= %s"); params.append(date_to)

    entries = []
    with _conn() as conn:
        # 1) 투자 정산 레그 — BUY/SELL/DIVIDEND 를 (날짜·계좌·종목·유형)으로 그룹
        ic, ip = ["t.type = ANY(%s)"], [["BUY", "SELL", "DIVIDEND"]]
        _filt(ic, ip)
        fills = conn.execute(
            f"""SELECT t.id, t.trade_date, t.type, t.symbol, t.name, t.currency,
                       t.quantity, t.price, t.amount, t.fee, t.tax,
                       a.id AS account_id, a.alias, a.account_no, o.name AS owner_name
                FROM transactions t JOIN accounts a ON a.id = t.account_id
                JOIN owners o ON o.id = a.owner_id
                WHERE {' AND '.join(ic)} ORDER BY t.trade_date DESC, t.id""", ip).fetchall()
        groups, order = {}, []
        for f in fills:
            key = (f["trade_date"], f["account_id"], f["symbol"], f["type"])
            g = groups.get(key)
            if g is None:
                g = groups[key] = {
                    "trade_date": f["trade_date"], "type": f["type"], "symbol": f["symbol"],
                    "name": f["name"], "currency": f["currency"], "owner_name": f["owner_name"],
                    "alias": f["alias"], "account_no": f["account_no"],
                    "quantity": 0.0, "_pv": 0.0, "amount": 0.0, "fee": 0.0, "tax": 0.0, "fills": []}
                order.append(key)
            g["quantity"] += f["quantity"] or 0
            g["_pv"] += (f["quantity"] or 0) * (f["price"] or 0)
            g["amount"] += f["amount"] or 0
            g["fee"] += f["fee"] or 0
            g["tax"] += f["tax"] or 0
            g["fills"].append({"id": f["id"], "quantity": f["quantity"], "price": f["price"],
                               "amount": f["amount"], "fee": f["fee"], "tax": f["tax"]})
        for key in order:
            g = groups[key]; t = g["type"]; gross = g["_pv"]
            if t == "BUY":
                cash, sign = -(gross + g["fee"] + g["tax"]), "neutral"
            elif t == "SELL":
                cash, sign = (gross - g["fee"] - g["tax"]), "neutral"
            else:  # DIVIDEND
                cash, sign = g["amount"], "income"
            entries.append({
                "trade_date": g["trade_date"], "owner_name": g["owner_name"], "alias": g["alias"],
                "account_no": g["account_no"], "type": t, "label": g["name"] or g["symbol"],
                "symbol": g["symbol"], "name": g["name"], "currency": g["currency"],
                "cash": cash, "sign": sign, "grouped": True, "fills": g["fills"],
                "quantity": g["quantity"], "avg_price": (gross / g["quantity"] if g["quantity"] else None),
                "fee": g["fee"], "tax": g["tax"],
                "_sortid": max((f["id"] for f in g["fills"]), default=0)})
        # 2) 순수 현금거래 — 모두 단건. 이체·환전·공모는 방향별 유형(중립).
        cc, cp = ["t.type = ANY(%s)"], [list(_CASHFLOW)]
        _filt(cc, cp)
        crows = conn.execute(
            f"""SELECT t.id, t.trade_date, t.type, t.name, t.currency, t.amount,
                       a.id AS account_id, a.alias, a.account_no, o.name AS owner_name
                FROM transactions t JOIN accounts a ON a.id = t.account_id
                JOIN owners o ON o.id = a.owner_id
                WHERE {' AND '.join(cc)} ORDER BY t.trade_date DESC, t.id""", cp).fetchall()
        for r in crows:
            t = r["type"]
            if t in _CASH_INCOME:
                cash, sign = r["amount"], "income"
            elif t in _CASH_EXPENSE:  # WITHDRAWAL, FEE, TAX
                cash, sign = -r["amount"], "expense"
            else:  # 이체/환전/공모 = 실제 입출금(중립 아님): 들어옴(in,+) / 나감(out,−)
                out = t.endswith("_OUT") or (not t.endswith("_IN") and "출금" in (r["name"] or ""))
                d, sign = (-1, "out") if out else (1, "in")
                cash = d * (r["amount"] or 0)
            entries.append({
                "trade_date": r["trade_date"], "owner_name": r["owner_name"], "alias": r["alias"],
                "account_no": r["account_no"], "type": t, "label": r["name"], "currency": r["currency"],
                "cash": cash, "sign": sign, "grouped": False, "id": r["id"], "_sortid": r["id"]})
    if type:
        entries = [e for e in entries if e["type"] == type]
    # 날짜 정렬 시 동일 날짜 안에서도 id(입력순) 기준으로 같은 방향 정렬
    keyfn = {"trade_date": lambda e: (e["trade_date"], e.get("_sortid", 0)),
             "cash": lambda e: abs(e["cash"] or 0),
             "type": lambda e: e["type"], "owner": lambda e: e["owner_name"] or ""}.get(
                 sort, lambda e: (e["trade_date"], e.get("_sortid", 0)))
    entries.sort(key=keyfn, reverse=(dir != "asc"))
    return {"total": len(entries), "entries": entries[offset:offset + limit], "limit": limit, "offset": offset}


# ── 이중기입 통합 원장 (P1) ──
@app.get("/api/movements")
def api_movements(
    owner: Optional[str] = None,
    account_id: Optional[str] = None,
    kind: Optional[str] = None,
    broker: Optional[str] = None,
    hide_kind: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: str = "trade_date",
    dir: str = "desc",
    limit: int = 100,
    offset: int = 0,
):
    """통합 거래(out→in): 한 행에 나감/들어옴 상품·수량·계좌."""
    clauses, params = [], []
    owners, accts = _multi(owner), _multi_int(account_id)
    if owners:
        clauses.append("(oo.name = ANY(%s) OR io.name = ANY(%s))"); params += [owners, owners]
    if accts:
        clauses.append("(m.out_account_id = ANY(%s) OR m.in_account_id = ANY(%s))"); params += [accts, accts]
    if kind:
        clauses.append("m.kind = %s"); params.append(kind)
    if broker:   # 증권사(내부 코드) 다중 선택
        bs = _multi(broker)
        clauses.append("(oa.brokerage = ANY(%s) OR ia.brokerage = ANY(%s))"); params += [bs, bs]
    if hide_kind:   # CMA 일일 스윕(예치·인출)처럼 노이즈인 유형 감추기 — 콤마로 여러 개
        clauses.append("m.kind <> ALL(%s)"); params.append(_multi(hide_kind))
    if q:
        clauses.append("(op.symbol ILIKE %s OR op.name ILIKE %s OR ip.symbol ILIKE %s OR ip.name ILIKE %s)")
        params += [f"%{q}%"] * 4
    if date_from:
        clauses.append("m.trade_date >= %s"); params.append(date_from)
    if date_to:
        clauses.append("m.trade_date <= %s"); params.append(date_to)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    joins = """FROM movements m
        LEFT JOIN accounts oa ON oa.id = m.out_account_id LEFT JOIN owners oo ON oo.id = oa.owner_id
        LEFT JOIN products op ON op.id = m.out_product_id
        LEFT JOIN accounts ia ON ia.id = m.in_account_id LEFT JOIN owners io ON io.id = ia.owner_id
        LEFT JOIN products ip ON ip.id = m.in_product_id"""
    with _conn() as conn:
        rows = conn.execute(
            f"""SELECT m.id, m.trade_date, m.kind, m.out_qty, m.in_qty, m.fee, m.tax, m.origin, m.adjustments, m.seq,
                       m.out_account_id, m.out_product_id, m.in_account_id, m.in_product_id,
                       oa.alias AS out_alias, oa.account_no AS out_acctno, oa.brokerage AS out_broker, oo.name AS out_owner,
                       op.symbol AS out_sym, op.name AS out_name, op.category AS out_cat, op.currency AS out_ccy, op.ticker AS out_pt,
                       ia.alias AS in_alias, ia.account_no AS in_acctno, ia.brokerage AS in_broker, io.name AS in_owner,
                       ip.symbol AS in_sym, ip.name AS in_name, ip.category AS in_cat, ip.currency AS in_ccy, ip.ticker AS in_pt
                {joins} {where} ORDER BY m.trade_date DESC, m.id DESC""", params).fetchall()
        single_acc = accts[0] if accts and len(accts) == 1 else None   # 상단 요약 패널용(행별 잔액과 별개)
    # (날짜·유형·나감계좌/상품·들어옴계좌/상품)으로 묶고, 체결별 수수료·세금은 펼쳐서 확인
    groups, order = {}, []
    for r in rows:
        key = (r["trade_date"], r["kind"], r["out_account_id"], r["out_product_id"],
               r["in_account_id"], r["in_product_id"])
        g = groups.get(key)
        if g is None:
            g = groups[key] = {k: r[k] for k in (
                "trade_date", "kind", "out_account_id", "in_account_id",
                "out_alias", "out_acctno", "out_broker", "out_owner", "out_sym", "out_name", "out_cat", "out_ccy", "out_pt",
                "in_alias", "in_acctno", "in_broker", "in_owner", "in_sym", "in_name", "in_cat", "in_ccy", "in_pt")}
            g.update(out_qty=0.0, in_qty=0.0, fills=[], _maxid=0, _seq=10 ** 9, adj={})
            order.append(key)
        g["out_qty"] += r["out_qty"] or 0
        g["in_qty"] += r["in_qty"] or 0
        g["_maxid"] = max(g["_maxid"], r["id"])
        g["_seq"] = min(g["_seq"], r["seq"] or 0)
        try:
            fadj = json.loads(r["adjustments"] or "[]")
        except Exception:
            fadj = []
        for a in fadj:      # 그룹 조정 합계(명목별)
            g["adj"][a["label"]] = g["adj"].get(a["label"], 0) + (a["amount"] or 0)
        g["fills"].append({"id": r["id"], "out_qty": r["out_qty"], "in_qty": r["in_qty"],
                           "adjustments": fadj, "origin": r["origin"], "seq": r["seq"] or 0})
    rev = (dir != "asc")
    out = [groups[k] for k in order]
    for g in out:   # 증권 티커 + 세부(체결)도 정렬 방향에 맞춤
        g["fills"].sort(key=lambda f: (f["seq"], f["id"]), reverse=rev)
        if g.get("out_cat") == "equity" and g.get("out_sym"):
            g["out_ticker"] = g.get("out_pt") or valuation._ticker_market(g["out_sym"])[0]
        if g.get("in_cat") == "equity" and g.get("in_sym"):
            g["in_ticker"] = g.get("in_pt") or valuation._ticker_market(g["in_sym"])[0]
    # 정렬: (날짜, seq, id) 한 방향. asc면 낮은 seq가 위, desc면 낮은 seq가 아래(날짜처럼 방향 따라 뒤집힘).
    key = (lambda g: (g["kind"], g["_seq"], g["_maxid"])) if sort == "kind" \
        else (lambda g: (g["trade_date"], g["_seq"], g["_maxid"]))
    out.sort(key=key, reverse=rev)
    groups = out[offset:offset + limit]
    # 행별 잔액: 그 거래 시점의 '계좌 누적 예수금'을 통화별로 항상 보여준다.
    # (현금이 안 움직인 입고·출고·배당 행에도 직전 잔액이 그대로 이어져 보이게 — 통장처럼)
    # 페이지에 등장하는 계좌만 running_cash 계산(계좌 단위 캐시) → 전체 건수와 무관하게 저비용.
    need = {a for g in groups for a in (g["out_account_id"], g["in_account_id"]) if a}
    if need:
        with _conn() as conn:
            rc = {a: movements.running_cash(conn, a) for a in need}
        # 그 계좌가 한 번이라도 쓴 통화만 열에 올린다(달러 안 쓰는 계좌에 0.00 도배 방지)
        seen = {a: {c for v in m.values() for c in v} for a, m in rc.items()}
        for g in groups:
            last = max(g["fills"], key=lambda f: (f["seq"], f["id"]))["id"]
            accs = [a for a in dict.fromkeys([g["out_account_id"], g["in_account_id"]]) if a]
            g["bal"] = [{"account_id": a,
                         **{c: (rc.get(a, {}).get(last, {}).get(c, 0.0) if c in seen.get(a, ()) else None)
                            for c in ("KRW", "USD")}} for a in accs]
    return {"total": len(out), "groups": groups, "limit": limit, "offset": offset,
            "single_account": single_acc}


class MovementIn(BaseModel):
    trade_date: str
    kind: str
    out_account_id: Optional[int] = None
    out_category: Optional[str] = None   # cash | equity
    out_symbol: Optional[str] = None     # 현금: KRW/USD, 증권: 종목명
    out_currency: Optional[str] = None
    out_ticker: Optional[str] = None
    out_qty: float = 0
    in_account_id: Optional[int] = None
    in_category: Optional[str] = None
    in_symbol: Optional[str] = None
    in_currency: Optional[str] = None
    in_ticker: Optional[str] = None
    in_qty: float = 0
    adjustments: Optional[list] = None   # [{"label":"수수료","amount":1000}, {"label":"할인","amount":-500}]
    note: Optional[str] = None


@app.post("/api/movements")
def api_movements_add(item: MovementIn):
    """수동 이중기입 거래 추가(예: 환전 KRW↔USD 한 번에)."""
    with _conn() as conn:
        db.init_schema(conn)
        return movements.add_manual(conn, item.dict())


@app.patch("/api/movements/{mid}")
def api_movements_update(mid: int, item: MovementIn):
    """이중기입 거래 수정(변환거래는 수동 승격)."""
    with _conn() as conn:
        db.init_schema(conn)
        return movements.update_movement(conn, mid, item.dict())


@app.delete("/api/movements/{mid}")
def api_movements_delete(mid: int):
    with _conn() as conn:
        db.init_schema(conn)
        return movements.delete_movement(conn, mid)


class MergeIn(BaseModel):
    out_id: int   # 현금 나감(출금) movement
    in_id: int    # 현금 들어옴(입금) movement


@app.post("/api/movements/merge")
def api_movements_merge(m: MergeIn):
    """서로 다른 계좌의 출금·입금 두 거래를 한 줄 '이체'로 합침."""
    with _conn() as conn:
        db.init_schema(conn)
        return movements.merge_transfer(conn, m.out_id, m.in_id)


@app.get("/api/movements/cash")
def api_movements_cash(owner: Optional[str] = None, account_id: Optional[str] = None,
                       broker: Optional[str] = None):
    """선택(소유자·계좌) 범위의 '현재' 현금 합계 = 예수금 + 예금성(RP·CMA·MMF).
    거래내역 필터바 한 줄 요약용 — 행별 잔액은 그 거래 시점 값이라 현재값과 다르다."""
    owners, accts, brokers = _multi(owner), _multi_int(account_id), _multi(broker)
    tot, n = {}, 0
    with _conn() as conn:
        for a in ledger.all_accounts(conn):
            if owners and a["owner_name"] not in owners:
                continue
            if accts and a["id"] not in accts:
                continue
            if brokers and a["brokerage"] not in brokers:
                continue
            n += 1
            for src in (movements.cash_by_ccy(conn, a["id"]), movements.deposits_by_ccy(conn, a["id"])):
                for c, v in src.items():
                    tot[c] = tot.get(c, 0.0) + v
    return {"n": n, "krw": round(tot.get("KRW", 0.0)), "usd": round(tot.get("USD", 0.0), 2)}


@app.get("/api/movements/daily")
def api_movements_daily(account_id: int, dates: str = ""):
    """계좌의 날짜별 '마감 잔액'(현금 통화별 + 종목 주수별). dates=쉼표구분."""
    ds = [d for d in dates.split(",") if d]
    with _conn() as conn:
        bal = movements.daily_balances(conn, account_id, ds)
        for b in bal.values():
            b["holdings"] = [{"name": h["name"], "ticker": h.get("ticker") or valuation._ticker_market(h["symbol"])[0],
                              "qty": h["qty"], "ccy": h["ccy"]} for h in b["holdings"]]
    return {"account_id": account_id, "balances": bal}


@app.get("/api/movements/asof")
def api_movements_asof(trade_date: str, max_id: int, account_id: str, seq: int = 0):
    """선택한 거래 시점까지의 계좌별 누적 잔액(현금 통화별 + 종목 주수별). 모달용. seq=표시순서."""
    accts = _multi_int(account_id)
    out = []
    with _conn() as conn:
        accs = {a["id"]: a for a in ledger.all_accounts(conn)}
        for aid in accts:
            a = accs.get(aid)
            if not a:
                continue
            b = movements.balances_as_of(conn, aid, trade_date, seq, max_id)
            holds = [{"name": h["name"], "ticker": h.get("ticker") or valuation._ticker_market(h["symbol"])[0],
                      "qty": h["qty"], "ccy": h["ccy"]} for h in b["holdings"]]
            out.append({"account_id": aid, "owner_name": a["owner_name"], "brokerage": a["brokerage"],
                        "alias": a["alias"], "account_no": a["account_no"], "cash": b["cash"], "holdings": holds})
    return {"accounts": out, "trade_date": trade_date}


@app.get("/api/movements/portfolio")
def api_movements_portfolio():
    """movements 기반 포지션·평가(대시보드 P2 검증용)."""
    with _conn() as conn:
        return valuation.portfolio(conn, positions_fn=movements.build_positions)


@app.post("/api/movements/rebuild")
def api_movements_rebuild():
    """transactions → movements 재생성(P1 전환용)."""
    with _conn() as conn:
        return movements.rebuild_movements(conn)


class ReorderIn(BaseModel):
    ids: list = []


# ── 시장 데이터(차트·배당·기업정보) + 관심종목 ─────────────────
# 화면은 DB만 읽는다. 야후는 하루 한 번 크론(cli.py market-refresh)이 부른다.
@app.get("/api/market/candles")
def api_market_candles(ticker: str, market: str = "", range: str = "1y"):
    with _conn() as conn:
        return market_mod.candles(conn, ticker, market, range)


@app.get("/api/market/profile")
def api_market_profile(ticker: str, market: str = ""):
    with _conn() as conn:
        return market_mod.profile(conn, ticker, market)


@app.post("/api/market/refresh")
def api_market_refresh(request: Request, all: int = 0):
    """관심·보유 종목의 시세·배당·기업정보를 지금 받아 온다(관리자)."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        return market_mod.refresh(conn, only_stale=not all)


class WatchStockIn(BaseModel):
    ticker: str
    name: str
    market: Optional[str] = None
    currency: Optional[str] = "KRW"
    target_krw: Optional[float] = None
    memo: Optional[str] = None
    group_id: Optional[int] = None


class WatchGroupIn(BaseModel):
    name: str


def _held_tickers(conn):
    """지금 보유 중인 종목(수량 > 0). 팔아서 0이 된 것은 빠진다."""
    out = {}
    for b in movements.pnl_by_symbol(conn, None):
        if b["qty"] > 1e-9 and (b.get("ticker") or "").strip():
            out[b["ticker"].upper()] = b
    return out


@app.get("/api/watch/stocks")
def api_watch_stocks():
    """그룹별 관심종목. 맨 앞은 거래내역에서 만든 '보유 중' 그룹(사람이 관리하지 않는다).
    시세는 전부 DB(symbol_candles)에서 읽으므로 종목이 많아도 외부 호출이 없다."""
    with _conn() as conn:
        def deco(ticker, base):
            last, chg = market_mod.last_close(conn, ticker)
            m = conn.execute(
                "SELECT sector, dividend_yield, ex_dividend, name FROM symbol_meta WHERE ticker = %s",
                (ticker,)).fetchone()
            r = dict(base)
            r["price"] = last
            r["change_pct"] = chg
            r["sector"] = m["sector"] if m else None
            r["dividend_yield"] = m["dividend_yield"] if m else None
            r["ex_dividend"] = m["ex_dividend"] if m else None
            if r.get("target_krw") and last:
                r["to_target_pct"] = (r["target_krw"] - last) / last * 100
            else:
                r["to_target_pct"] = None
            return r

        held = _held_tickers(conn)
        groups = [{
            "id": "held", "name": "보유 중", "virtual": True,
            "rows": [deco(t, {"ticker": t, "name": b["name"], "market": b["market"],
                              "currency": b["ccy"], "qty": b["qty"], "target_krw": None})
                     for t, b in sorted(held.items(), key=lambda kv: -(kv[1]["qty"] or 0))],
        }]
        for g in conn.execute("SELECT id, name FROM watch_groups ORDER BY sort_order, id").fetchall():
            rows = [dict(r) for r in conn.execute(
                "SELECT * FROM watch_stocks WHERE group_id = %s ORDER BY created_at DESC, id DESC",
                (g["id"],)).fetchall()]
            for r in rows:
                r["created_at"] = str(r["created_at"])[:10]
            groups.append({"id": g["id"], "name": g["name"], "virtual": False,
                           "rows": [deco(r["ticker"], r) for r in rows]})
    return {"groups": groups}


@app.post("/api/watch/groups")
def api_watch_group_add(item: WatchGroupIn):
    name = (item.name or "").strip()
    if not name:
        return {"error": "그룹 이름을 입력하세요."}
    with _conn() as conn:
        row = conn.execute(
            """INSERT INTO watch_groups(name, sort_order)
               VALUES (%s, COALESCE((SELECT max(sort_order) + 1 FROM watch_groups), 0))
               ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name RETURNING id""",
            (name,)).fetchone()
        conn.commit()
    return {"ok": True, "id": row["id"]}


@app.delete("/api/watch/groups/{gid}")
def api_watch_group_del(gid: int):
    """그룹을 지우면 그 안의 종목도 함께 빠진다(다른 그룹에 담긴 같은 종목은 남는다)."""
    with _conn() as conn:
        conn.execute("DELETE FROM watch_stocks WHERE group_id = %s", (gid,))
        conn.execute("DELETE FROM watch_groups WHERE id = %s", (gid,))
        conn.commit()
    return {"ok": True}


@app.post("/api/watch/stocks")
def api_watch_stock_add(item: WatchStockIn):
    tk = (item.ticker or "").strip().upper()
    if not tk:
        return {"error": "티커가 필요합니다."}
    with _conn() as conn:
        gid = item.group_id or (conn.execute(
            "SELECT id FROM watch_groups ORDER BY sort_order, id LIMIT 1").fetchone() or {}).get("id")
        if not gid:
            gid = conn.execute("INSERT INTO watch_groups(name) VALUES ('관심') RETURNING id").fetchone()["id"]
        row = conn.execute(
            """INSERT INTO watch_stocks(ticker, name, market, currency, target_krw, memo, group_id)
               VALUES (%s,%s,%s,%s,%s,%s,%s)
               ON CONFLICT (group_id, ticker) DO UPDATE SET name = EXCLUDED.name,
                 market = EXCLUDED.market, currency = EXCLUDED.currency
               RETURNING id""",
            (tk, (item.name or tk).strip(), item.market, (item.currency or "KRW").upper(),
             item.target_krw, item.memo, gid)).fetchone()
        conn.commit()
        # 처음 담은 종목이면 시세를 그 자리에서 한 번 채운다(빈 차트를 보여 줄 수 없다).
        if not conn.execute("SELECT 1 FROM symbol_candles WHERE ticker = %s LIMIT 1", (tk,)).fetchone():
            market_mod.refresh_one(conn, tk, item.market or "", full=True)
    return {"ok": True, "id": row["id"], "group_id": gid}


@app.patch("/api/watch/stocks/{wid}")
def api_watch_stock_edit(wid: int, item: WatchStockIn):
    with _conn() as conn:
        conn.execute("UPDATE watch_stocks SET target_krw = %s, memo = %s WHERE id = %s",
                     (item.target_krw, item.memo, wid))
        if item.group_id:
            conn.execute("UPDATE watch_stocks SET group_id = %s WHERE id = %s", (item.group_id, wid))
        conn.commit()
    return {"ok": True}


@app.delete("/api/watch/stocks/{wid}")
def api_watch_stock_del(wid: int):
    with _conn() as conn:
        conn.execute("DELETE FROM watch_stocks WHERE id = %s", (wid,))
        conn.commit()
    return {"ok": True}


@app.get("/api/symbols/search")
def api_symbols_search(q: str = "", limit: int = 10, market: str = ""):
    """종목명/티커 자동완성 — 보유종목 + symbols.csv + 전체 상장목록. market=kr/us 로 국내·미국 필터."""
    ql = q.strip().lower()
    if not ql:
        return {"items": []}
    from .instruments import normalize_name
    want_ccy = {"kr": "KRW", "us": "USD"}.get(market)   # None = 전체
    seen, items = set(), []

    def add(name, ticker, ccy, mkt):
        cy = (ccy or "KRW")
        if want_ccy and cy != want_ccy:
            return
        # 티커 있으면 티커로, 없으면 정규화명으로 중복 제거(철자 변형·보유 쪼개짐 방지)
        key = "T:" + ticker.upper() if ticker else "N:" + normalize_name(name or "")
        if not name or key in seen:
            return
        seen.add(key)
        items.append({"name": name, "ticker": ticker or "", "ccy": cy, "market": mkt or ""})

    like = f"%{ql}%"
    with _conn() as conn:
        # 1) 보유종목(내 데이터, 한글명 우선)
        for r in conn.execute(
            """SELECT DISTINCT symbol, name, ticker, currency, market FROM products
               WHERE category='equity' AND (lower(symbol) LIKE %s OR lower(name) LIKE %s OR lower(ticker) LIKE %s)
               ORDER BY symbol LIMIT 20""", (like, like, like)):
            tk = r["ticker"] or valuation._ticker_market(r["symbol"])[0]
            add(r["symbol"] or r["name"], tk, r["currency"], r["market"])
        # 2) 큐레이션 매핑(symbols.csv — 미국종목 한글명 등)
        import csv as _csv
        if config.SYMBOLS_CSV.exists():
            for row in _csv.DictReader(open(config.SYMBOLS_CSV, encoding="utf-8-sig")):
                nm, tk = (row.get("normalized_name") or "").strip(), (row.get("ticker") or "").strip()
                if nm and (ql in nm.lower() or ql in tk.lower()):
                    ccy = "USD" if tk.isalpha() and tk.isupper() else "KRW"
                    add(nm, tk, ccy, valuation._MARKETS.get(tk, ""))
        # 3) 전체 상장목록(symbols 테이블 — 티커 정확일치 우선, 짧은 이름 우선)
        for r in conn.execute(
            """SELECT ticker, name, market, currency FROM symbols
               WHERE (lower(name) LIKE %s OR lower(ticker) LIKE %s)
                 AND (%s = '' OR currency = %s)
               ORDER BY (lower(ticker) = %s) DESC, char_length(name) LIMIT 40""",
            (like, like, (want_ccy or ""), (want_ccy or ""), ql)):
            add(r["name"], r["ticker"], r["currency"], r["market"])
    return {"items": items[:limit]}


@app.post("/api/symbols/sync")
def api_symbols_sync(request: Request):
    """전체 상장종목(KRX + 미국) FDR로 수집해 symbols 테이블 갱신. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from . import symbols as symbols_mod
    with _conn() as conn:
        return symbols_mod.sync_symbols(conn)


class AliasIn(BaseModel):
    name: str
    ticker: str
    currency: str = ""
    market: str = ""


@app.post("/api/symbols/alias")
def api_symbol_alias(item: AliasIn, request: Request):
    """증권사 한글명 → 티커 별칭 등록(재시작 불필요, ≤30초 내 반영). 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from .instruments import normalize_name
    key = normalize_name(item.name or "")
    tk = (item.ticker or "").strip().upper()
    if not key or not tk:
        return {"error": "종목명·티커를 입력하세요."}
    ccy = (item.currency or "").strip().upper() or ("USD" if tk.isalpha() else "KRW")
    with _conn() as conn:
        db.init_schema(conn)
        conn.execute(
            """INSERT INTO symbol_aliases(name, ticker, market, currency) VALUES (%s,%s,%s,%s)
               ON CONFLICT (name) DO UPDATE SET ticker=EXCLUDED.ticker, market=EXCLUDED.market, currency=EXCLUDED.currency""",
            (key, tk, (item.market or "").strip(), ccy))
        conn.commit()
    valuation.reload_symmap()
    return {"ok": True, "name": key, "ticker": tk, "currency": ccy}


@app.delete("/api/symbols/alias")
def api_symbol_alias_delete(name: str, request: Request):
    """별칭 삭제. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from .instruments import normalize_name
    with _conn() as conn:
        conn.execute("DELETE FROM symbol_aliases WHERE name = %s", (normalize_name(name or ""),))
        conn.commit()
    valuation.reload_symmap()
    return {"ok": True}


@app.get("/api/symbols/aliases")
def api_symbol_aliases():
    """등록된 별칭 목록 + 전체 상장목록 수."""
    with _conn() as conn:
        db.init_schema(conn)
        aliases = [dict(r) for r in conn.execute(
            "SELECT name, ticker, market, currency FROM symbol_aliases ORDER BY name")]
        n = conn.execute("SELECT count(*) c FROM symbols").fetchone()["c"]
        disp = {r["skey"]: r["display"] for r in conn.execute("SELECT skey, display FROM symbol_display")}
        # 거래내역에 등장한 모든 주식(보유 여부 무관 — 매도해 보유0이어도 포함)
        prods = conn.execute(
            """SELECT DISTINCT p.name, p.symbol, p.ticker FROM products p
               WHERE p.category = 'equity' AND EXISTS (
                 SELECT 1 FROM movements m WHERE m.out_product_id = p.id OR m.in_product_id = p.id)
               ORDER BY p.name""").fetchall()
    from .instruments import is_cash_equivalent
    # 티커 기준으로 묶어 목록 축소(같은 종목이 증권사별 이름으로 중복 → 한 줄). 티커 없으면 이름별.
    groups = {}
    for r in prods:
        nm = r["name"] or r["symbol"]
        if is_cash_equivalent(nm):   # RP/MMF/CMA 등 현금성은 종목 관리 대상 아님
            continue
        tk = (r["ticker"] or "")
        if not tk:
            try:
                tk = valuation._ticker_market(nm)[0] or ""
            except Exception:
                tk = ""
        key = tk if tk else "\x00" + nm   # 티커 없으면 이름별 고유
        g = groups.setdefault(key, {"ticker": tk, "names": []})
        if nm not in g["names"]:
            g["names"].append(nm)
    instruments = sorted(groups.values(), key=lambda g: g["names"][0])
    return {"aliases": aliases, "symbols_count": n, "display": disp, "instruments": instruments}


@app.get("/api/symbols/display")
def api_symbols_display():
    """종목 표시명(별칭) 맵. 모든 사용자 렌더에 적용되므로 공개."""
    with _conn() as conn:
        db.init_schema(conn)
        rows = conn.execute("SELECT skey, display FROM symbol_display").fetchall()
    return {"display": {r["skey"]: r["display"] for r in rows}}


class DisplayIn(BaseModel):
    key: str                 # 티커(우선) 또는 원본 종목명
    display: str = ""        # 빈 값이면 표시명 삭제(원래 이름으로)


@app.post("/api/symbols/display")
def api_symbols_display_set(item: DisplayIn, request: Request):
    """종목 표시명 등록/삭제. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    key = (item.key or "").strip()
    disp = (item.display or "").strip()
    if not key:
        return {"error": "종목 키가 없습니다."}
    with _conn() as conn:
        db.init_schema(conn)
        if disp:
            conn.execute("""INSERT INTO symbol_display(skey, display) VALUES (%s,%s)
                            ON CONFLICT (skey) DO UPDATE SET display=EXCLUDED.display""", (key, disp))
        else:
            conn.execute("DELETE FROM symbol_display WHERE skey = %s", (key,))
        conn.commit()
    return {"ok": True, "key": key, "display": disp}


@app.post("/api/admin/reset")
def api_admin_reset(request: Request):
    """거래내역 초기화(거래·movements·상품·계좌·예수금·스냅샷 삭제). 로그인·시세는 유지. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        for t in ("movements", "transactions", "cash_balances", "snapshots", "accounts", "products"):
            conn.execute(f"DELETE FROM {t}")
        conn.commit()
    return {"ok": True}


_BROKER_KR = {"kb": "KB증권", "mirae": "미래에셋증권", "kiwoom": "키움증권", "samsung": "삼성증권"}


def _build_export_workbook(conn):
    """저장 데이터 → 내보내기 양식 워크북(메타·계좌·거래내역·원본거래) + 데이터 해시.
    내보내기(다운로드)·저장(서버) 공용. 반환 (wb, data_hash, now_str)."""
    import hashlib

    import openpyxl

    def _adj_cols(js, n=5):
        try:
            arr = [a for a in json.loads(js or "[]") if a.get("amount")]
        except Exception:
            arr = []
        cells = []
        for k in range(n):
            if k < len(arr):
                a = arr[k]
                ccy = (a.get("ccy") or "").upper()
                amt = a.get("amount")
                cells += [a.get("label", ""), f"{amt} {ccy}" if ccy and ccy != "KRW" else amt]
            else:
                cells += ["", ""]
        return cells

    adj_hdr = [x for k in range(1, 6) for x in (f"조정{k}", f"조정{k}금액")]
    wb = openpyxl.Workbook()
    # 1) 계좌 요약
    ws = wb.active
    ws.title = "계좌"
    ws.append(["소유자", "증권사", "계좌번호", "계좌명", "거래수", "원화잔액", "외화(USD)잔액"])
    accts = conn.execute(
        """SELECT a.id, o.name owner, a.brokerage, a.account_no, a.alias,
                  (SELECT count(*) FROM transactions t WHERE t.account_id = a.id) txc
           FROM accounts a JOIN owners o ON o.id = a.owner_id
           ORDER BY o.name, a.brokerage, a.account_no""").fetchall()
    for a in accts:
        cash = movements.cash_by_ccy(conn, a["id"])
        ws.append([a["owner"], _BROKER_KR.get(a["brokerage"], a["brokerage"]), a["account_no"],
                   a["alias"] or "", a["txc"], round(cash.get("KRW", 0)), round(cash.get("USD", 0), 2)])
    # 2) 거래내역(movements) — 행마다 통화별 시점 잔액
    rc_cache = {}

    def _bal(acc_id, mid):
        if acc_id not in rc_cache:
            rc_cache[acc_id] = movements.running_cash(conn, acc_id)
        b = rc_cache[acc_id].get(mid, {})
        return b.get("KRW", ""), b.get("USD", "")

    wm = wb.create_sheet("거래내역")
    wm.append(["소유자", "계좌", "날짜", "유형", "나감상품", "나감수량", "들어옴상품", "들어옴수량", "취득원가",
               "원화잔액", "외화(USD)잔액", *adj_hdr, "비고"])
    for r in conn.execute(
        """SELECT o.name owner, a.alias, a.account_no, a.id acc_id, m.trade_date, m.kind, m.seq, m.id, m.cost,
                  op.name o_name, op.symbol o_sym, m.out_qty,
                  ip.name i_name, ip.symbol i_sym, m.in_qty, m.adjustments, m.note
           FROM movements m
           LEFT JOIN accounts a ON a.id = COALESCE(m.out_account_id, m.in_account_id)
           LEFT JOIN owners o ON o.id = a.owner_id
           LEFT JOIN products op ON op.id = m.out_product_id
           LEFT JOIN products ip ON ip.id = m.in_product_id
           ORDER BY o.name, a.account_no, m.trade_date, m.seq, m.id""").fetchall():
        bk, bu = _bal(r["acc_id"], r["id"])
        wm.append([r["owner"] or "", f"{r['alias'] or ''}({r['account_no'] or ''})", str(r["trade_date"]),
                   r["kind"], r["o_name"] or r["o_sym"] or "", r["out_qty"] or "",
                   r["i_name"] or r["i_sym"] or "", r["in_qty"] or "", r["cost"] or "", bk, bu,
                   *_adj_cols(r["adjustments"]), r["note"] or ""])
    # 3) 원본거래(transactions) — 동시에 dedupe_hash 수집(데이터 지문)
    wt = wb.create_sheet("원본거래")
    wt.append(["소유자", "계좌", "날짜", "유형", "종목", "수량", "단가", "금액", "수수료", "세금", "통화", "출처"])
    tx_hashes = []
    for t in conn.execute(
        """SELECT o.name owner, a.alias, a.account_no, t.trade_date, t.type, t.symbol, t.name,
                  t.quantity, t.price, t.amount, t.fee, t.tax, t.currency, t.source, t.dedupe_hash
           FROM transactions t JOIN accounts a ON a.id = t.account_id JOIN owners o ON o.id = a.owner_id
           ORDER BY o.name, a.account_no, t.trade_date, t.id""").fetchall():
        wt.append([t["owner"], f"{t['alias'] or ''}({t['account_no'] or ''})", str(t["trade_date"]), t["type"],
                   t["name"] or t["symbol"] or "", t["quantity"] or "", t["price"] or "", t["amount"] or "",
                   t["fee"] or "", t["tax"] or "", t["currency"], t["source"] or ""])
        if t["dedupe_hash"]:
            tx_hashes.append(t["dedupe_hash"])
    # 데이터 해시: 전 거래 dedupe_hash 정렬 결합의 SHA256 → 무결성·중복확인(실데이터 저장 대비)
    data_hash = hashlib.sha256("|".join(sorted(tx_hashes)).encode("utf-8")).hexdigest()
    now = conn.execute("SELECT to_char(now(),'YYYY-MM-DD HH24:MI:SS') t").fetchone()["t"]
    nmv = conn.execute("SELECT count(*) c FROM movements").fetchone()["c"]
    # 4) 메타(맨 앞 시트)
    meta = wb.create_sheet("메타", 0)
    meta.append(["생성일시", now])
    meta.append(["데이터 해시(SHA256)", data_hash])
    meta.append(["계좌수", len(accts)])
    meta.append(["거래수", len(tx_hashes)])
    meta.append(["movement수", nmv])
    meta.append(["해시 산식", "전 거래 dedupe_hash 정렬 결합의 SHA256 (무결성·중복확인용)"])
    return wb, data_hash, now


@app.get("/api/export.xlsx")
def api_export_xlsx(request: Request):
    """저장된 데이터를 내보내기 양식 xlsx로 다운로드(메타·계좌·거래내역·원본거래). 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    import io
    with _conn() as conn:
        wb, data_hash, _ = _build_export_workbook(conn)
        today = conn.execute("SELECT CURRENT_DATE d").fetchone()["d"]
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="goraes_export_{today}_{data_hash[:8]}.xlsx"'})


@app.post("/api/export/save")
def api_export_save(request: Request):
    """같은 양식 xlsx를 서버에 보관(data/exports/saved/). 파일명·시트에 데이터 해시 포함. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        wb, data_hash, now = _build_export_workbook(conn)
    outdir = config.EXPORTS_DIR   # 공유폴더 거래내역/export (파일앱·시놀로지 드라이브에서 접근)
    outdir.mkdir(parents=True, exist_ok=True)
    stamp = now.replace("-", "").replace(":", "").replace(" ", "_")
    fname = f"goraes_{stamp}_{data_hash[:8]}.xlsx"
    wb.save(str(outdir / fname))
    return {"ok": True, "file": fname, "hash": data_hash, "saved_at": now, "dir": str(outdir)}


@app.get("/api/stock")
def api_stock(symbol: str, owner: Optional[str] = None):
    """종목 상세: 거래이력(매수·매도·입고·출고·배당) + 보유수량·평균취득가·실현손익(native)."""
    with _conn() as conn:
        pids = [r["id"] for r in conn.execute(
            "SELECT id FROM products WHERE category = 'equity' AND (symbol = %s OR name = %s)",
            (symbol, symbol)).fetchall()]
        if not pids:
            return {"symbol": symbol, "trades": [], "qty": 0, "avg_cost": 0, "realized": 0, "currency": "KRW"}
        rows = conn.execute(
            """SELECT m.trade_date, m.kind, m.out_qty, m.in_qty, m.cost, m.adjustments,
                      m.out_product_id o_pid, m.in_product_id i_pid,
                      op.currency o_ccy, ip.currency i_ccy,
                      COALESCE(oa.alias, oa.account_no, ia.alias, ia.account_no) acct,
                      oo.name o_owner, io.name i_owner
               FROM movements m
               LEFT JOIN products op ON op.id = m.out_product_id
               LEFT JOIN products ip ON ip.id = m.in_product_id
               LEFT JOIN accounts oa ON oa.id = m.out_account_id LEFT JOIN owners oo ON oo.id = oa.owner_id
               LEFT JOIN accounts ia ON ia.id = m.in_account_id LEFT JOIN owners io ON io.id = ia.owner_id
               WHERE m.out_product_id = ANY(%s) OR m.in_product_id = ANY(%s)
               ORDER BY m.trade_date, m.seq, m.id""", (pids, pids)).fetchall()
        qty = cost = realized = 0.0
        ccy = "KRW"
        trades = []
        for m in rows:
            if owner and owner not in (m["o_owner"], m["i_owner"]):
                continue
            k = m["kind"]
            t = {"date": str(m["trade_date"]), "kind": k, "account": m["acct"] or "",
                 "qty": 0, "price": 0, "cash": 0, "adj": _adj_text(m["adjustments"])}
            if m["i_pid"] in pids and k in ("매수", "입고"):
                q = m["in_qty"] or 0
                c = (m["out_qty"] or 0) if k == "매수" else (m["cost"] or 0)
                qty += q; cost += c; ccy = m["i_ccy"] or ccy
                t.update(qty=q, cash=c, price=(c / q if q else 0))
            elif m["o_pid"] in pids and k in ("매도", "출고"):
                q = min(m["out_qty"] or 0, qty)
                unit = (cost / qty) if qty > 1e-9 else 0
                if k == "매도":
                    realized += (m["in_qty"] or 0) - unit * q
                qty -= q; cost -= unit * q; ccy = m["o_ccy"] or ccy
                t.update(qty=-(m["out_qty"] or 0), cash=(m["in_qty"] or 0),
                         price=((m["in_qty"] or 0) / (m["out_qty"] or 1)))
            elif k == "배당" and m["o_pid"] in pids:
                t.update(cash=(m["in_qty"] or 0))
            else:
                continue
            trades.append(t)
    return {"symbol": symbol, "currency": ccy, "qty": round(qty, 4),
            "avg_cost": round(cost / qty, 2) if qty > 1e-9 else 0,
            "realized": round(realized, 2), "trades": trades}


def _adj_text(js):
    try:
        arr = [a for a in json.loads(js or "[]") if a.get("amount")]
    except Exception:
        return ""
    return " · ".join(f"{a.get('label', '')} {a.get('amount')}{(a.get('ccy') or '')}" for a in arr)


@app.get("/api/reconcile")
def api_reconcile(request: Request):
    """데이터 검증(대사): 계좌별 계산 현금(movements) vs 브로커 예수금 스냅샷(cash_balances)의 차이. 관리자 전용."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    rows = []
    with _conn() as conn:
        for a in conn.execute(
            """SELECT a.id, o.name owner, a.brokerage, a.account_no, a.alias
               FROM accounts a JOIN owners o ON o.id = a.owner_id
               ORDER BY o.name, a.brokerage, a.account_no""").fetchall():
            comp = movements.cash_by_ccy(conn, a["id"])
            brk, asof = {}, None
            for r in conn.execute("SELECT currency, balance, as_of FROM cash_balances WHERE account_id = %s",
                                  (a["id"],)).fetchall():
                brk[r["currency"]] = r["balance"] or 0
                if r["as_of"] and (asof is None or str(r["as_of"]) > str(asof)):
                    asof = r["as_of"]

            def cell(ccy):
                c, b = comp.get(ccy, 0), brk.get(ccy, 0)
                return {"computed": round(c, 2), "broker": round(b, 2), "diff": round(c - b, 2)}

            rows.append({
                "owner": a["owner"], "brokerage": _BROKER_KR.get(a["brokerage"], a["brokerage"]),
                "account": f"{a['alias'] or ''}({a['account_no']})",
                "krw": cell("KRW"), "usd": cell("USD"), "as_of": str(asof) if asof else "",
            })
    return {"rows": rows}


@app.post("/api/movements/reorder")
def api_movements_reorder(body: ReorderIn):
    """날짜 내 수동 정렬: 넘어온 movement id 순서(위→아래)대로 seq=0..n 부여."""
    with _conn() as conn:
        for i, mid in enumerate(body.ids):
            conn.execute("UPDATE movements SET seq=%s WHERE id=%s", (i, int(mid)))
        conn.commit()
    return {"ok": True, "n": len(body.ids)}


@app.get("/api/pnl-symbols")
def api_pnl_symbols(owner: str = ""):
    """사고판 종목별 손익 — 지금 안 들고 있어도 나온다.
    실현손익은 판 만큼, 평가손익은 남은 만큼. 원화 환산은 현재 환율 기준이다
    (매도 시점 환율이 아니므로 해외 종목은 참고값)."""
    from .prices import base as prices
    owners = [x for x in owner.split(",") if x.strip()]
    with _conn() as conn:
        rows = movements.pnl_by_symbol(conn, owners or None)
        out = []
        for b in rows:
            ccy = b["ccy"]
            fx = prices.get_fx(conn, ccy) or 1.0
            px = prices.get_price(conn, b["symbol"]) if b["qty"] > 1e-9 else None
            mv = (px * b["qty"]) if px else 0.0
            unreal = (mv - b["cost"]) if px else 0.0
            total = b["realized"] + unreal + b["dividend"]
            out.append({
                **{k: b[k] for k in ("symbol", "name", "market", "ccy", "ticker",
                                     "qty", "buy_qty", "buy_amt", "sell_qty", "sell_amt",
                                     "realized", "dividend", "first", "last")},
                "cost": b["cost"], "price": px, "market_value": mv,
                "unrealized": unreal, "total": total, "fx": fx,
                "realized_krw": b["realized"] * fx, "dividend_krw": b["dividend"] * fx,
                "unrealized_krw": unreal * fx, "total_krw": total * fx,
                "held": b["qty"] > 1e-9,
            })
    out.sort(key=lambda r: r["total_krw"], reverse=True)
    agg = lambda k: sum(r[k] for r in out)
    return {"rows": out, "total": {
        "realized_krw": agg("realized_krw"), "dividend_krw": agg("dividend_krw"),
        "unrealized_krw": agg("unrealized_krw"), "total_krw": agg("total_krw"),
        "n": len(out), "n_held": sum(1 for r in out if r["held"])}}


@app.get("/api/fx")
def api_fx(owner: str = "", account_id: str = ""):
    """환전 내역과 '내가 산 환율'. 매입 평균은 주식 평단과 같은 이동평균이다 —
    되팔면 수량만 줄고 평균은 그대로다(그래야 남은 달러의 취득단가가 맞다)."""
    where, params = ["m.kind = '환전'"], []
    ids = [int(v) for v in account_id.split(",") if v.strip().isdigit()]
    if ids:
        where.append("(m.out_account_id = ANY(%s) OR m.in_account_id = ANY(%s))")
        params += [ids, ids]
    if owner:
        where.append("o.name = ANY(%s)"); params.append([x for x in owner.split(",") if x])
    with _conn() as conn:
        rows = conn.execute(
            f"""SELECT m.id, m.trade_date, m.out_qty, m.in_qty,
                       po.symbol AS out_sym, pi.symbol AS in_sym,
                       a.id AS account_id, a.brokerage, a.alias, a.account_no, o.name AS owner
                FROM movements m
                JOIN products po ON po.id = m.out_product_id
                JOIN products pi ON pi.id = m.in_product_id
                JOIN accounts a ON a.id = COALESCE(m.out_account_id, m.in_account_id)
                LEFT JOIN owners o ON o.id = a.owner_id
                WHERE {' AND '.join(where)}
                ORDER BY m.trade_date, m.id""", params).fetchall()
        fxrow = conn.execute("SELECT price FROM prices WHERE price_key = %s",
                             ("FX:USDKRW",)).fetchone()
    now_fx = float(fxrow["price"]) if fxrow else 0.0

    out, pos = [], {}          # pos[통화] = {"qty": 보유, "cost": 원화 취득총액}
    for r in rows:
        buy = (r["out_sym"] == "KRW")                     # 원화 → 외화
        ccy = r["in_sym"] if buy else r["out_sym"]
        krw = (r["out_qty"] if buy else r["in_qty"]) or 0
        fx = (r["in_qty"] if buy else r["out_qty"]) or 0
        rate = (krw / fx) if fx else 0
        p = pos.setdefault(ccy, {"qty": 0.0, "cost": 0.0, "buy_krw": 0.0, "buy_fx": 0.0,
                                 "sell_fx": 0.0, "sell_krw": 0.0, "realized": 0.0})
        if buy:
            p["qty"] += fx; p["cost"] += krw
            p["buy_fx"] += fx; p["buy_krw"] += krw
        else:
            avg = (p["cost"] / p["qty"]) if p["qty"] else 0
            used = min(fx, p["qty"])
            p["realized"] += krw - used * avg              # 환차익(되판 만큼)
            p["qty"] -= used; p["cost"] -= used * avg
            p["sell_fx"] += fx; p["sell_krw"] += krw
        out.append({
            "id": r["id"], "trade_date": str(r["trade_date"]), "side": "매수" if buy else "매도",
            "ccy": ccy, "krw": krw, "fx": fx, "rate": rate,
            "account_id": r["account_id"], "owner": r["owner"],
            "brokerage": r["brokerage"], "alias": r["alias"], "account_no": r["account_no"],
            "avg_after": (p["cost"] / p["qty"]) if p["qty"] else 0,
        })

    summary = []
    for ccy, p in pos.items():
        # '순매수'는 환전으로 조달한 달러에서 되판 만큼을 뺀 것이다. 그 달러로 주식을 샀으면
        # 현금엔 안 남아 있으니 '보유'라고 부르면 안 된다(평가손익도 그래서 계산하지 않는다).
        avg = (p["cost"] / p["qty"]) if p["qty"] > 1e-9 else 0
        summary.append({
            "ccy": ccy,
            "net_fx": p["qty"], "net_krw": p["cost"], "avg_rate": avg,
            "buy_fx": p["buy_fx"], "buy_krw": p["buy_krw"],
            "buy_avg": (p["buy_krw"] / p["buy_fx"]) if p["buy_fx"] else 0,
            "sell_fx": p["sell_fx"], "sell_krw": p["sell_krw"],
            "sell_avg": (p["sell_krw"] / p["sell_fx"]) if p["sell_fx"] else 0,
            "realized_krw": p["realized"],           # 되판 달러에서 실제로 생긴 환차익
            "now_fx": now_fx,
            "gap": (now_fx - avg) if avg else 0,      # 지금 환율이 내 평균보다 얼마나 위인가
            "gap_pct": ((now_fx / avg - 1) * 100) if avg else 0,
        })
    out.reverse()              # 화면은 최신순
    return {"rows": out, "summary": summary, "now_fx": now_fx}


@app.get("/api/movements/meta")
def api_movements_meta(account_id: str = ""):
    """account_id(쉼표 구분)를 주면 유형 목록을 그 계좌들에 실제로 있는 것만 준다.
    필터가 앞에서부터 좁혀지는데 유형만 전체가 뜨면, 고르고 나서 빈 화면을 보게 된다."""
    ids = [int(v) for v in account_id.split(",") if v.strip().lstrip("-").isdigit()]
    with _conn() as conn:
        owners = [r["name"] for r in conn.execute("SELECT name FROM owners ORDER BY name")]
        accounts = [dict(a) for a in ledger.all_accounts(conn)]
        if ids:
            kinds = [r["kind"] for r in conn.execute(
                """SELECT DISTINCT kind FROM movements
                   WHERE out_account_id = ANY(%s) OR in_account_id = ANY(%s) ORDER BY kind""",
                (ids, ids))]
        else:
            kinds = [r["kind"] for r in conn.execute("SELECT DISTINCT kind FROM movements ORDER BY kind")]
        months = [r["ym"] for r in conn.execute(
            "SELECT DISTINCT to_char(trade_date::date, 'YYYY-MM') AS ym FROM movements ORDER BY ym DESC")]
        brokers = [r["brokerage"] for r in conn.execute(
            "SELECT DISTINCT brokerage FROM accounts WHERE brokerage IS NOT NULL ORDER BY brokerage")]
    return {"owners": owners, "accounts": accounts, "kinds": kinds, "months": months, "brokers": brokers}




class AccountIn(BaseModel):
    brokerage: str
    account_no: str
    alias: Optional[str] = None


@app.post("/api/account")
def api_account_add(item: AccountIn, request: Request):
    """수동 계좌 추가(소유자=로그인 사용자). 이후 거래를 직접 입력할 수 있음."""
    from .ingest.adapters import resolve_brokerage
    user = _current_user(request)
    if not user:
        return {"error": "로그인이 필요합니다."}
    owner = (user.get("owner") or user.get("name") or "").strip()
    if not owner:
        return {"error": "소유자 정보가 없습니다. 관리자에게 소유자 지정을 요청하세요."}
    acct = (item.account_no or "").strip()
    if not acct:
        return {"error": "계좌번호를 입력하세요."}
    key = resolve_brokerage(item.brokerage or "") or (item.brokerage or "").strip()
    with _conn() as conn:
        db.init_schema(conn)
        oid = db.get_or_create_owner(conn, owner)
        aid = db.get_or_create_account(conn, oid, key, acct, (item.alias or "").strip() or None)
        conn.commit()
    return {"id": aid, "owner": owner, "brokerage": key, "account_no": acct}


class AccountEditIn(BaseModel):
    account_no: str
    alias: Optional[str] = None
    brokerage: Optional[str] = None


@app.patch("/api/account/{aid}")
def api_account_edit(aid: int, item: AccountEditIn, request: Request):
    """등록 계좌 수정(계좌번호·계좌명·증권사). 관리자 전용. 다른 계좌와 (증권사,계좌번호) 중복 방지."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from .ingest.adapters import resolve_brokerage
    acct = (item.account_no or "").strip()
    if not acct:
        return {"error": "계좌번호를 입력하세요."}
    alias = (item.alias or "").strip() or None
    with _conn() as conn:
        row = conn.execute("SELECT brokerage FROM accounts WHERE id = %s", (aid,)).fetchone()
        if not row:
            return {"error": "계좌를 찾을 수 없어요."}
        brk = resolve_brokerage(item.brokerage or "") or (item.brokerage or "").strip() or row["brokerage"]
        dup = conn.execute("SELECT id FROM accounts WHERE brokerage = %s AND account_no = %s AND id <> %s",
                           (brk, acct, aid)).fetchone()
        if dup:
            return {"error": f"이미 있는 계좌예요 ({brk} {acct})."}
        conn.execute("UPDATE accounts SET account_no = %s, alias = %s, brokerage = %s WHERE id = %s",
                     (acct, alias, brk, aid))
        conn.commit()
    return {"ok": True, "id": aid, "account_no": acct, "alias": alias, "brokerage": brk}


@app.post("/api/upload")
async def api_upload(
    request: Request,
    file: UploadFile = File(...),
    brokerage: str = Form(...),
    account_no: str = Form(...),
    alias: str = Form(""),
):
    """로그인 사용자를 소유자로, 올린 증권사 CSV를 (증권사·계좌번호·계좌명) 지정해 적재."""
    import tempfile
    from .ingest.adapters import check_format, parse_stats, resolve_brokerage
    from .ingest.importer import import_file

    user = _current_user(request)
    if not user:
        return {"error": "로그인이 필요합니다."}
    owner = (user.get("owner") or user.get("name") or "").strip()
    if not owner:
        return {"error": "소유자 정보가 없습니다. 관리자에게 가족 소유자 지정을 요청하세요."}
    account_no = (account_no or "").strip()
    key = resolve_brokerage(brokerage or "")
    if not account_no:
        return {"error": "계좌번호를 입력하세요."}
    if not key:
        return {"error": f"알 수 없는 증권사: {brokerage}"}

    data = await file.read()
    suffix = Path(file.filename or "upload.csv").suffix or ".csv"
    tmp = tempfile.NamedTemporaryFile(prefix="upload_", suffix=suffix, delete=False)
    try:
        tmp.write(data); tmp.close()
        try:  # 진단용: 마지막 업로드 원본 항상 보존
            import shutil
            shutil.copy(tmp.name, config.DATA_DIR / "_last_upload.csv")
        except Exception:
            pass
        # 형식 검증 먼저 — 안 맞으면 1건도 넣지 않고 거부
        ok, header, missing = check_format(tmp.name, key)
        if not ok:
            return {"error": f"{brokerage} 형식과 맞지 않아요. 누락 열: {', '.join(missing)}",
                    "missing": missing, "header": header, "file": file.filename}
        # 적재율 검증 — 형식은 맞아도 인식 거래가 너무 적으면(수량·금액 비어있는 리포트 등) 거부
        data_rows, yielded = parse_stats(tmp.name, key)
        if data_rows >= 4 and yielded < max(2, data_rows * 0.3):
            return {"error": f"형식은 맞지만 인식된 거래가 {yielded}/{data_rows}건뿐이에요. "
                             f"수량·금액이 비어있거나 다른 리포트(예: 원화정산 내역)일 수 있어요.",
                    "yielded": yielded, "rows": data_rows, "file": file.filename}
        with _conn() as conn:
            db.init_schema(conn)
            res = import_file(conn, tmp.name, owner, key, account_no, (alias or "").strip() or None)
            if res.get("inserted"):
                movements.rebuild_movements(conn)   # 통합 원장 동기화
        res.update(file=file.filename, owner=owner, brokerage=key, account_no=account_no)
        # imports 폴더에도 표준 이름으로 보관 → 웹/폴더 어느 쪽으로 넣어도 imports가 최신 유지(NAS 원천).
        if res.get("inserted") or res.get("skipped"):
            try:
                import hashlib
                import shutil
                from .ingest.adapters import (canonical_import_name, parse_file,
                                              year_span)
                yr = year_span(list(parse_file(tmp.name, key))) or hashlib.sha1(data).hexdigest()[:6]
                ext = Path(file.filename or "u.csv").suffix or ".csv"
                rel = canonical_import_name(owner, (alias or "").strip(), key, account_no, yr, ext)
                dest = config.IMPORTS_DIR / rel
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy(tmp.name, dest)
                res["saved_to"] = rel
            except Exception:
                pass
        if not res.get("inserted") and not res.get("skipped"):
            res["diag"] = _upload_diag(tmp.name)  # 0건이면 인식 헤더/인코딩 진단
            try:  # 진단용 원본 보존(0건일 때만)
                import shutil
                shutil.copy(tmp.name, config.DATA_DIR / "_last_upload.csv")
            except Exception:
                pass
        return res
    except Exception as e:
        return {"error": str(e), "file": file.filename}
    finally:
        os.unlink(tmp.name)


def _upload_diag(path):
    """0건 파싱 시 원인 진단: 인코딩별 첫 줄(헤더) 미리보기."""
    import csv
    out = {}
    for enc in ("cp949", "utf-8-sig"):
        try:
            with open(path, newline="", encoding=enc) as f:
                cols = next(csv.reader(f), [])
            out[enc] = [c.strip() for c in cols][:12]
        except Exception as e:
            out[enc] = f"decode 실패: {e}"
    return out


@app.post("/api/imports/scan")
def api_imports_scan(request: Request):
    """imports 폴더 즉시 스캔(관리자). 크론(매분)과 동일 로직 — advisory lock으로 중복 방지."""
    if not _require_admin(request):
        return {"error": "관리자만 실행할 수 있어요."}
    from .ingest.importer import scan_imports
    with _conn() as conn:
        db.init_schema(conn)
        return scan_imports(conn, config.IMPORTS_DIR)


@app.post("/api/refresh-prices")
def api_refresh_prices(request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from .prices import fdr
    with _conn() as conn:
        return fdr.refresh(conn)


# ---------------- 자산추이·보유자산·배당·거시지표 ----------------
@app.get("/api/snapshots")
def api_snapshots(scope: str = "TOTAL"):
    with _conn() as conn:
        rows = conn.execute(
            "SELECT as_of, market_value_krw, cash_krw, realestate_krw, total_krw "
            "FROM snapshots WHERE scope = %s ORDER BY as_of", (scope,)).fetchall()
    return rows


def _owner_list(owners: str | None):
    """'김영한,김숙진' → ['김영한','김숙진']. 비었으면 None(=전체)."""
    if not owners:
        return None
    v = [o.strip() for o in owners.split(",") if o.strip()]
    return v or None


@app.get("/api/nav-monthly")
def api_nav_monthly(owners: str = None):
    """자산의 월별 추이(전 기간). 스냅샷은 일별로도 쌓이므로 각 달의 '마지막 날'만 골라 월말로 본다.
    소유자를 주면 그 소유자들의 scope를 합산한다(백필이 소유자별 scope를 이미 저장해 둔다).
    비어 있는 달은 직전 달 값을 이어 그린다 — 선이 끊기면 자산이 0이 된 것처럼 보인다."""
    who = _owner_list(owners)
    scopes = who if who else ["TOTAL"]
    with _conn() as conn:
        rows = conn.execute(
            """SELECT substring(as_of,1,7) AS month, scope,
                      market_value_krw, cash_krw, realestate_krw, total_krw
               FROM (
                 SELECT DISTINCT ON (scope, substring(as_of,1,7)) *
                 FROM snapshots WHERE scope = ANY(%s)
                 ORDER BY scope, substring(as_of,1,7), as_of DESC
               ) x ORDER BY month""", (scopes,)).fetchall()
    if not rows:
        return []
    KEYS = ("market_value_krw", "cash_krw", "realestate_krw", "total_krw")
    per_month, last_by_scope = {}, {}
    for r in rows:
        m = r["month"]
        per_month.setdefault(m, {})[r["scope"]] = {k: (r[k] or 0) for k in KEYS}

    months = sorted(per_month)
    # 첫 달 ~ 마지막 달을 빠짐없이 만든다(중간에 비는 달이 있어도 축이 고르게 흐르도록)
    def step(m):
        y, mo = int(m[:4]), int(m[5:7])
        return f"{y + 1}-01" if mo == 12 else f"{y}-{mo + 1:02d}"
    full, cur = [], months[0]
    while cur <= months[-1]:
        full.append(cur)
        cur = step(cur)

    out = []
    for m in full:
        for sc, v in per_month.get(m, {}).items():
            last_by_scope[sc] = v                        # 그 달에 값이 온 scope만 갱신
        if not last_by_scope:
            continue                                     # 아직 아무 scope도 시작 안 함
        agg = {k: 0.0 for k in KEYS}
        for v in last_by_scope.values():                 # 나머지는 직전 값을 이어 쓴다
            for k in KEYS:
                agg[k] += v[k]
        out.append({"month": m, **{k: round(agg[k]) for k in KEYS}})
    return out


@app.get("/api/income-monthly")
def api_income_monthly(owners: str = None):
    """배당·이자의 월별 흐름(전 기간, 원화환산). 세금은 원천징수(transactions.tax).
    환율은 현재 환율을 쓴다(기존 배당 집계와 같은 기준)."""
    who = _owner_list(owners)
    fx = "COALESCE(p.price, 1)"
    sql = f"""SELECT substring(t.trade_date,1,7) AS month,
                sum(CASE WHEN t.type='DIVIDEND' THEN t.amount * {fx} ELSE 0 END) AS div_gross,
                sum(CASE WHEN t.type='DIVIDEND' THEN COALESCE(t.tax,0) * {fx} ELSE 0 END) AS div_tax,
                sum(CASE WHEN t.type='INTEREST' THEN t.amount * {fx} ELSE 0 END) AS int_gross,
                sum(CASE WHEN t.type='INTEREST' THEN COALESCE(t.tax,0) * {fx} ELSE 0 END) AS int_tax
              FROM transactions t
              JOIN accounts a ON a.id = t.account_id
              JOIN owners o ON o.id = a.owner_id
              LEFT JOIN prices p ON t.currency <> 'KRW' AND p.price_key = 'FX:' || t.currency || 'KRW'
              WHERE t.type IN ('DIVIDEND','INTEREST') {"AND o.name = ANY(%s)" if who else ""}
              GROUP BY month ORDER BY month"""
    with _conn() as conn:
        rows = conn.execute(sql, (who,) if who else ()).fetchall()
    out = []
    for r in rows:
        dg, dt = r["div_gross"] or 0, r["div_tax"] or 0
        ig, it = r["int_gross"] or 0, r["int_tax"] or 0
        out.append({"month": r["month"],
                    "div": round(dg), "div_tax": round(dt), "div_net": round(dg - dt),
                    "int": round(ig), "int_tax": round(it), "int_net": round(ig - it),
                    "net": round(dg - dt + ig - it)})
    return out


@app.post("/api/snapshot")
def api_snapshot(request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        db.init_schema(conn)
        return valuation.save_snapshot(conn)


@app.post("/api/snapshots/backfill")
def api_snapshots_backfill(request: Request):
    """최초 거래월부터 이번 달까지 각 월말 순자산을 과거 시세로 역산해 채움(관리자, FDR 조회로 수십 초 소요 가능)."""
    if not _require_admin(request):
        return JSONResponse({"error": "관리자만 실행할 수 있어요."}, status_code=403)
    with _conn() as conn:
        db.init_schema(conn)
        return valuation.backfill_monthly_snapshots(conn)


class OwnedIn(BaseModel):
    name: str
    owner: Optional[str] = None
    category: Optional[str] = "부동산"
    kind: Optional[str] = "자가"          # 자가/전세/월세/임대/대출/기타자산/기타부채
    value_krw: int = 0                    # 현재 시세/보증금/잔액/금액
    loan_krw: int = 0                     # 자가 대출(net 차감)
    monthly_krw: int = 0                  # 월세(메모)
    as_of: Optional[str] = None           # 현재값 기준일 YYYY-MM-DD
    acquire_date: Optional[str] = None    # 취득일/계약시작
    acquire_krw: int = 0                  # 취득가/최초보증금
    dispose_date: Optional[str] = None    # 매도일/종료일(이후 제외)
    dispose_krw: int = 0                  # 매도가/반환액(= dis_p1..p4 합계, 서버가 계산)
    # 대금은 단계별로 받는다. 합계가 취득가/매도가가 되고,
    # 전세·월세는 이 합계가 곧 보증금(value_krw)이다.
    acq_p1: int = 0                       # 가계약금
    acq_p2: int = 0                       # 계약금
    acq_p3: int = 0                       # 중도금
    acq_p4: int = 0                       # 잔금
    dis_p1: int = 0
    dis_p2: int = 0
    dis_p3: int = 0
    dis_p4: int = 0
    loan_ids: List[int] = []              # 이 부동산에 걸린 대출(부채 항목 id들)
    note: Optional[str] = None
    re_sgg: Optional[str] = None          # 실거래가 연결 — 자치구
    re_apt: Optional[str] = None          # 실거래가 연결 — 단지명
    re_area: Optional[float] = None       # 실거래가 연결 — 전용면적(㎡)
    link_owned_id: Optional[int] = None   # 이 부채가 걸린 실물자산(주담대 → 집)
    link_account_id: Optional[int] = None # 이 부채가 걸린 계좌(마이너스통장 등)


@app.get("/api/owned-assets")
def api_owned_assets(history: int = 0):
    """history=0: 현재 보유중 항목 + net + 순액 합계(자산−부채). history=1: 전 항목(매도 포함)."""
    from datetime import date
    today = date.today().isoformat()
    with _conn() as conn:
        db.init_schema(conn)
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM owned_assets ORDER BY COALESCE(owner,''), name, acquire_date NULLS FIRST, id").fetchall()]
    by_owned = {}
    for r in rows:                          # 어떤 부채가 어느 부동산에 걸렸는지 미리 모은다
        if r.get("link_owned_id"):
            by_owned.setdefault(r["link_owned_id"], []).append(
                {"id": r["id"], "name": r["name"], "value_krw": r["value_krw"] or 0})
    for r in rows:
        r["loans"] = by_owned.get(r["id"], [])
        r["loan_linked_krw"] = sum(x["value_krw"] for x in r["loans"])
        r["net_krw"] = valuation.owned_net(r.get("kind"), r.get("value_krw"), r.get("loan_krw"))
        held = not (r.get("dispose_date") and r["dispose_date"] <= today)
        if r.get("acquire_date") and r["acquire_date"] > today:
            held = False
        r["held"] = held
    if history:
        return {"items": rows}
    held = [r for r in rows if r["held"]]
    held.sort(key=lambda r: r["net_krw"], reverse=True)
    return {"items": held, "total": sum(r["net_krw"] for r in held)}


DEPOSIT_KINDS = ("전세", "월세", "임대")     # 보증금이 곧 값인 것들


def _owned_norm(item: OwnedIn):
    """단계별 대금을 합쳐 취득가·매도가를 만든다. 전세·월세는 그 합계가 곧 보증금.
    대금을 한 칸도 안 적었으면 기존에 직접 친 값을 그대로 둔다(옛 데이터 보호)."""
    acq_parts = (item.acq_p1, item.acq_p2, item.acq_p3, item.acq_p4)
    dis_parts = (item.dis_p1, item.dis_p2, item.dis_p3, item.dis_p4)
    acq = sum(acq_parts) or item.acquire_krw
    dis = sum(dis_parts) or item.dispose_krw
    val = item.value_krw
    if item.kind in DEPOSIT_KINDS and sum(acq_parts):
        val = acq                          # 보증금 = 낸 돈의 합
    if item.dispose_date and dis:
        val = 0                            # 판 것에는 시세가 없다. 매도가가 그 자리를 대신한다.
    return acq, dis, val


def _owned_cols(item: OwnedIn):
    acq, dis, val = _owned_norm(item)
    # 담보대출을 부채 목록에서 고르면 loan_krw는 쓰지 않는다.
    # 그 대출은 부채 항목으로 이미 순액에서 −로 잡히므로, 여기서 또 빼면 이중 차감이다.
    loan = 0 if item.loan_ids else item.loan_krw
    return (item.owner, item.category, item.kind, item.name, val, loan,
            item.monthly_krw, item.as_of, item.acquire_date, acq,
            item.dispose_date, dis, item.note,
            item.re_sgg, item.re_apt, item.re_area,
            item.link_owned_id, item.link_account_id,
            item.acq_p1, item.acq_p2, item.acq_p3, item.acq_p4,
            item.dis_p1, item.dis_p2, item.dis_p3, item.dis_p4)


def _relink_loans(conn, owned_id: int, loan_ids):
    """이 부동산에 걸린 대출을 다시 맨다. 뺀 것은 연결만 풀고 항목은 남긴다."""
    conn.execute("UPDATE owned_assets SET link_owned_id = NULL "
                 "WHERE link_owned_id = %s AND NOT (id = ANY(%s))", (owned_id, loan_ids or [0]))
    if loan_ids:
        conn.execute("UPDATE owned_assets SET link_owned_id = %s WHERE id = ANY(%s)",
                     (owned_id, loan_ids))


@app.post("/api/owned-assets")
def api_owned_add(item: OwnedIn, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        db.init_schema(conn)
        row = conn.execute(
            """INSERT INTO owned_assets(owner, category, kind, name, value_krw, loan_krw, monthly_krw,
                   as_of, acquire_date, acquire_krw, dispose_date, dispose_krw, note,
                   re_sgg, re_apt, re_area, link_owned_id, link_account_id,
                   acq_p1, acq_p2, acq_p3, acq_p4, dis_p1, dis_p2, dis_p3, dis_p4)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            _owned_cols(item)).fetchone()
        _relink_loans(conn, row["id"], item.loan_ids)
        conn.commit()
    return {"id": row["id"]}


@app.patch("/api/owned-assets/{item_id}")
def api_owned_edit(item_id: int, item: OwnedIn, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        db.init_schema(conn)
        conn.execute(
            """UPDATE owned_assets SET owner=%s, category=%s, kind=%s, name=%s, value_krw=%s, loan_krw=%s,
                   monthly_krw=%s, as_of=%s, acquire_date=%s, acquire_krw=%s, dispose_date=%s, dispose_krw=%s,
                   note=%s, re_sgg=%s, re_apt=%s, re_area=%s,
                   link_owned_id=%s, link_account_id=%s,
                   acq_p1=%s, acq_p2=%s, acq_p3=%s, acq_p4=%s,
                   dis_p1=%s, dis_p2=%s, dis_p3=%s, dis_p4=%s
               WHERE id=%s""",
            (*_owned_cols(item), item_id))
        _relink_loans(conn, item_id, item.loan_ids)
        conn.commit()
    return {"ok": True}


@app.delete("/api/owned-assets/{item_id}")
def api_owned_del(item_id: int, request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    with _conn() as conn:
        conn.execute("DELETE FROM owned_assets WHERE id = %s", (item_id,))
        conn.commit()
    return {"deleted": item_id}


@app.get("/api/dividends-monthly")
def api_dividends_monthly():
    with _conn() as conn:
        rows = conn.execute(
            """SELECT substring(t.trade_date,1,7) AS month,
                      sum(t.amount * COALESCE(p.price,1)) AS krw
               FROM transactions t
               LEFT JOIN prices p ON t.currency <> 'KRW' AND p.price_key = 'FX:' || t.currency || 'KRW'
               WHERE t.type = 'DIVIDEND'
               GROUP BY month ORDER BY month""").fetchall()
    return rows


@app.get("/api/dividends-summary")
def api_dividends_summary():
    """배당 요약(원화환산): 연도별 총액·세금·순액, 종목별 순액. 세금=원천징수(transactions.tax)."""
    fxj = ("LEFT JOIN prices p ON t.currency <> 'KRW' AND p.price_key = 'FX:' || t.currency || 'KRW'")
    gross = "sum(t.amount * COALESCE(p.price, 1))"
    tax = "sum(COALESCE(t.tax,0) * COALESCE(p.price, 1))"
    with _conn() as conn:
        by_year = conn.execute(
            f"""SELECT substring(t.trade_date,1,4) yr, {gross} gross, {tax} tax
                FROM transactions t {fxj} WHERE t.type='DIVIDEND'
                GROUP BY yr ORDER BY yr""").fetchall()
        by_stock = conn.execute(
            f"""SELECT COALESCE(NULLIF(t.name,''), t.symbol) nm, {gross} gross, {tax} tax, count(*) n
                FROM transactions t {fxj} WHERE t.type='DIVIDEND'
                GROUP BY nm ORDER BY gross DESC""").fetchall()
    fmt = lambda rs, key: [{**{k: r[k] for k in r.keys()},
                            "net": round((r["gross"] or 0) - (r["tax"] or 0))} for r in rs]
    return {"by_year": fmt(by_year, "yr"), "by_stock": fmt(by_stock, "nm")}


@app.get("/api/macro")
def api_macro():
    with _conn() as conn:
        return conn.execute("SELECT * FROM macro ORDER BY category, name").fetchall()


@app.post("/api/macro-refresh")
def api_macro_refresh(request: Request):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    from . import macro
    with _conn() as conn:
        db.init_schema(conn)
        return macro.refresh(conn)


# ---------------- 부동산 (실거래가 + 관심매물) ----------------
@app.get("/api/re/meta")
def re_meta():
    with _conn() as conn:
        n = conn.execute("SELECT count(*) AS c FROM re_apt_trades").fetchone()["c"]
        last = conn.execute("SELECT max(deal_date) AS d FROM re_apt_trades").fetchone()["d"]
    return {
        "gu": [{"code": c, "name": n} for c, n in SEOUL_GU.items()],
        "trade_count": n, "last_deal_date": last,
        "has_key": bool(config.MOLIT_SERVICE_KEY),
    }


_RE_SORT = {
    "deal_date": "deal_date", "apt_name": "apt_name", "area": "area", "floor": "floor",
    "deal_amount": "deal_amount", "build_year": "build_year", "sgg": "sgg_name",
}


@app.get("/api/re/transactions")
def re_transactions(
    sgg: Optional[str] = None, apt: Optional[str] = None,
    area_min: Optional[float] = None, area_max: Optional[float] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    deal_type: Optional[str] = None,       # 매매/전세/월세 (쉼표 구분, 비우면 전부)
    sort: str = "deal_date", dir: str = "desc",
    limit: int = 100, offset: int = 0,
):
    clauses, params = [], []
    types = [t for t in (deal_type or "").split(",") if t.strip()]
    if types:
        clauses.append("t.deal_type = ANY(%s)"); params.append(types)
    if sgg:
        clauses.append("t.sgg_cd = %s"); params.append(sgg)
    if apt:
        clauses.append("t.apt_name ILIKE %s"); params.append(f"%{apt}%")
    if area_min is not None:
        clauses.append("t.area >= %s"); params.append(area_min)
    if area_max is not None:
        clauses.append("t.area <= %s"); params.append(area_max)
    if date_from:
        clauses.append("t.deal_date >= %s"); params.append(date_from)
    if date_to:
        clauses.append("t.deal_date <= %s"); params.append(date_to)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    order_col = _RE_SORT.get(sort, "deal_date")
    order_dir = "ASC" if dir == "asc" else "DESC"
    with _conn() as conn:
        total = conn.execute(f"SELECT count(*) AS c FROM re_apt_trades t {where}", params).fetchone()["c"]
        rows = conn.execute(
            f"""SELECT t.sgg_name, t.umd, t.apt_name, t.area, t.floor, t.deal_amount, t.deal_date,
                       t.deal_type, t.monthly_rent, t.build_year, t.road_name,
                       b.bc_rat, b.vl_rat, b.land_share
                FROM re_apt_trades t
                LEFT JOIN re_buildings b
                  ON b.sgg_cd = t.sgg_cd AND b.umd = t.umd AND b.jibun = t.jibun
                {where}
                ORDER BY t.{order_col} {order_dir}, t.id DESC LIMIT %s OFFSET %s""",
            params + [limit, offset],
        ).fetchall()
    return {"total": total, "rows": rows}


@app.get("/api/re/lookup")
def re_lookup(q: str = "", area_min: Optional[float] = None, area_max: Optional[float] = None,
              limit: int = 30):
    """실거래가에 있는 단지를 (구·단지·면적)별로 찾아 최근 거래를 붙여 준다.
    보유 부동산을 여기에 연결해 두면 시세를 직접 끌어올 수 있다."""
    if not (q or "").strip():
        return {"rows": []}
    clauses, params = ["apt_name ILIKE %s"], [f"%{q.strip()}%"]
    if area_min is not None:
        clauses.append("area >= %s"); params.append(area_min)
    if area_max is not None:
        clauses.append("area <= %s"); params.append(area_max)
    where = "WHERE " + " AND ".join(clauses)
    with _conn() as conn:
        # 한 단지·면적의 매매·전세·월세를 한 줄로 모아 준다 — 자가/전세/월세 어느 쪽을
        # 등록하든 그 종류에 맞는 시세를 바로 집어 쓸 수 있어야 한다.
        rows = conn.execute(
            f"""SELECT sgg_name, apt_name, round(area::numeric, 2) AS area,
                       count(*) AS deals, max(deal_date) AS last_date,
                       count(*) FILTER (WHERE deal_type='매매') AS n_sale,
                       count(*) FILTER (WHERE deal_type='전세') AS n_jeonse,
                       count(*) FILTER (WHERE deal_type='월세') AS n_rent,
                       (array_agg(deal_amount ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='매매'))[1] AS sale_amount,
                       (array_agg(deal_date ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='매매'))[1] AS sale_date,
                       (array_agg(deal_amount ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='전세'))[1] AS jeonse_amount,
                       (array_agg(deal_date ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='전세'))[1] AS jeonse_date,
                       (array_agg(deal_amount ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='월세'))[1] AS rent_deposit,
                       (array_agg(monthly_rent ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='월세'))[1] AS rent_monthly,
                       (array_agg(deal_date ORDER BY deal_date DESC)
                          FILTER (WHERE deal_type='월세'))[1] AS rent_date
                FROM re_apt_trades {where}
                GROUP BY sgg_name, apt_name, round(area::numeric, 2)
                ORDER BY last_date DESC, deals DESC LIMIT %s""",
            params + [limit]).fetchall()
    return {"rows": [dict(r) for r in rows]}


@app.get("/api/re/quote")
def re_quote(sgg: str = "", apt: str = "", area: Optional[float] = None,
             deal_type: str = "매매", n: int = 5):
    """연결된 단지의 최근 실거래 몇 건. 시세를 채울 때 뭘 보고 넣는지 알 수 있어야 한다."""
    clauses, params = ["apt_name = %s"], [apt]
    if deal_type:
        clauses.append("deal_type = %s"); params.append(deal_type)
    if sgg:
        clauses.append("sgg_name = %s"); params.append(sgg)
    if area is not None:
        clauses.append("abs(area - %s) < 0.5"); params.append(area)
    with _conn() as conn:
        rows = conn.execute(
            f"""SELECT deal_date, deal_amount, monthly_rent, deal_type, area, floor
                FROM re_apt_trades
                WHERE {" AND ".join(clauses)} ORDER BY deal_date DESC, id DESC LIMIT %s""",
            params + [n]).fetchall()
    rows = [dict(r) for r in rows]
    return {"rows": rows, "last": rows[0] if rows else None}


class WatchIn(BaseModel):
    apt_name: str
    sgg_name: Optional[str] = None
    area: Optional[float] = None
    floor: Optional[int] = None
    price: Optional[int] = None      # 호가(만원)
    url: Optional[str] = None
    note: Optional[str] = None


@app.get("/api/re/watchlist")
def re_watchlist():
    """등록 매물 + 각 단지(±면적) 실거래 시세 요약(건수·최저~최고·평균·㎡당·추이) 첨부."""
    from datetime import date, timedelta
    cut = (date.today() - timedelta(days=45)).isoformat()  # 최근/이전 추이 기준
    with _conn() as conn:
        return conn.execute(
            """SELECT l.id, l.sgg_name, l.apt_name, l.area, l.floor, l.price, l.url, l.note,
                      s.cnt, s.avg_amt, s.min_amt, s.max_amt, s.per_sqm,
                      s.last_deal, s.last_deal_date, s.last_floor,
                      s.recent_avg, s.older_avg,
                      bd.bc_rat, bd.vl_rat, bd.land_share, bd.hhld_cnt
               FROM re_listings l
               LEFT JOIN LATERAL (
                   SELECT count(*) AS cnt,
                          round(avg(deal_amount)) AS avg_amt,
                          min(deal_amount) AS min_amt,
                          max(deal_amount) AS max_amt,
                          round(avg(deal_amount / NULLIF(area,0))) AS per_sqm,
                          (array_agg(deal_amount ORDER BY deal_date DESC, id DESC))[1] AS last_deal,
                          (array_agg(deal_date  ORDER BY deal_date DESC, id DESC))[1] AS last_deal_date,
                          (array_agg(floor      ORDER BY deal_date DESC, id DESC))[1] AS last_floor,
                          round(avg(deal_amount) FILTER (WHERE deal_date >= %s)) AS recent_avg,
                          round(avg(deal_amount) FILTER (WHERE deal_date <  %s)) AS older_avg
                   FROM re_apt_trades t
                   WHERE t.deal_type = '매매' AND t.apt_name = l.apt_name
                     AND (l.area IS NULL OR abs(t.area - l.area) < 1.5)
               ) s ON true
               LEFT JOIN LATERAL (
                   SELECT bc_rat, vl_rat, land_share, hhld_cnt FROM re_buildings rb
                   WHERE rb.apt_name = l.apt_name
                   ORDER BY rb.hhld_cnt DESC NULLS LAST LIMIT 1
               ) bd ON true
               ORDER BY l.created_at DESC""",
            (cut, cut),
        ).fetchall()


@app.post("/api/re/watchlist")
def re_watchlist_add(item: WatchIn):
    with _conn() as conn:
        # 이미 담은 단지·면적이면 새로 만들지 않고 비어 있던 값만 채운다.
        row = conn.execute(
            """INSERT INTO re_listings (sgg_name, apt_name, area, floor, price, url, note)
               VALUES (%s,%s,%s,%s,%s,%s,%s)
               ON CONFLICT (apt_name, COALESCE(sgg_name,''), COALESCE(area, -1)) DO UPDATE SET
                 floor = COALESCE(EXCLUDED.floor, re_listings.floor),
                 price = COALESCE(EXCLUDED.price, re_listings.price),
                 url   = COALESCE(EXCLUDED.url,   re_listings.url),
                 note  = COALESCE(EXCLUDED.note,  re_listings.note)
               RETURNING id, (xmax <> 0) AS existed""",
            (item.sgg_name, item.apt_name, item.area, item.floor, item.price, item.url, item.note),
        ).fetchone()
        conn.commit()
    return {"id": row["id"], "existed": row["existed"]}


@app.delete("/api/re/watchlist/{item_id}")
def re_watchlist_del(item_id: int):
    with _conn() as conn:
        conn.execute("DELETE FROM re_listings WHERE id = %s", (item_id,))
        conn.commit()
    return {"deleted": item_id}


@app.post("/api/re/sync")
def re_sync(request: Request, months: int = 1, kinds: str = "매매,전월세"):
    if not _require_admin(request):
        return JSONResponse({"error": "관리자 전용"}, status_code=403)
    # 버튼 수동 갱신은 최근 1개월만(요청 타임아웃 이내). 전체 백필은 cron(cli re-sync)이 담당.
    from .realestate import molit
    ks = tuple(k for k in kinds.split(",") if k.strip()) or ("매매",)
    with _conn() as conn:
        db.init_schema(conn)
        return molit.sync_seoul(conn, months=months, kinds=ks)


